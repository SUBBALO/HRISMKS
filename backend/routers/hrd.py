"""HRD module: Master Karyawan, Slip Gaji (fleksibel), blast email via Gmail SMTP.
Portal dikunci PIN — bahkan admin harus masukkan PIN untuk melihat data gaji."""
import os
import io
import uuid
import hashlib
import hmac
import smtplib
import ssl
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.utils import formatdate, make_msgid

import jwt
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db import db
from deps import get_current_user, log_action, is_super_admin_user
from security import hash_password, verify_password, JWT_SECRET, JWT_ALGORITHM
from services.soft_delete import NOT_DELETED_FILTER, soft_delete_one, merged

router = APIRouter(prefix="/hrd", tags=["hrd"])

ROMAN = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X", 11: "XI", 12: "XII"}
BULAN_ID = {1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember"}

# Template pesan email default (bisa diubah dari menu Pengaturan Email)
DEFAULT_EMAIL_SUBJECT = "Slip Gaji {bulan} {tahun} - {nama}"
DEFAULT_EMAIL_BODY = (
    "Yth. {nama},\n\n"
    "Berikut kami lampirkan slip gaji Anda untuk periode {bulan} {tahun}.\n"
    "Take Home Pay: {take_home}.\n\n"
    "Dokumen ini bersifat rahasia. Mohon tidak menyebarkan.\n\n"
    "Hormat kami,\n{sender}"
)


class _SafeDict(dict):
    def __missing__(self, key):
        return "{" + key + "}"


def _smtp_friendly(err: str) -> str:
    """Ubah pesan error SMTP teknis menjadi panduan yang bisa ditindaklanjuti (umum, semua provider)."""
    e = err or ""
    low = e.lower()
    if "5.7.26" in e or "unauthenticated" in low or ("spf" in low and "dkim" in low):
        return ("Ditolak server penerima: domain pengirim belum lolos autentikasi SPF/DKIM. "
                "Aktifkan SPF & DKIM untuk domain Anda di pengaturan DNS hosting.")
    if "5.7.8" in e or "username and password not accepted" in low or "authenticationerror" in low or "authentication failed" in low:
        return "Login SMTP ditolak. Periksa Username & Password email di Pengaturan Email."
    if "authentication required" in low or "5.7.0" in e:
        return "Autentikasi diperlukan. Isi Username & Password SMTP di Pengaturan Email."
    if "5.1.1" in e or "does not exist" in low or "no such user" in low:
        return "Alamat email tujuan tidak ditemukan. Periksa kembali email karyawan."
    if "getaddrinfo" in low or "name or service not known" in low or "connection refused" in low or "timed out" in low or "timeout" in low:
        return "Gagal terhubung ke server SMTP. Periksa SMTP Host, Port, dan pilihan Keamanan (SSL/TLS)."
    if "wrong version number" in low or "ssl" in low:
        return "Kesalahan SSL/TLS. Untuk port 465 pilih SSL, untuk port 587 pilih TLS (STARTTLS)."
    return e[:220]


def _open_smtp(host: str, port: int, security: str, username: str, password: str, timeout: int = 30):
    """Buka koneksi SMTP sesuai mode keamanan lalu login. security: 'ssl' atau 'tls'."""
    context = ssl.create_default_context()
    if (security or "ssl").lower() == "ssl":
        server = smtplib.SMTP_SSL(host, port, context=context, timeout=timeout)
    else:  # tls / starttls
        server = smtplib.SMTP(host, port, timeout=timeout)
        server.ehlo()
        server.starttls(context=context)
        server.ehlo()
    server.login(username, password)
    return server

# HRD menus (Accurate-style granular permission). Each menu supports actions:
# view, create, edit, delete, report
HRD_MENUS = [
    {"key": "hrd_karyawan", "label": "Master Karyawan", "group": "gaji"},
    {"key": "hrd_slip_gaji", "label": "Slip Gaji", "group": "gaji"},
    {"key": "hrd_email", "label": "Kirim Email Slip", "group": "gaji"},
    {"key": "hrd_settings", "label": "Pengaturan Email", "group": "gaji"},
    {"key": "hrd_dokumen", "label": "Dokumen HRD", "group": "dokumen"},
]
HRD_MENU_KEYS = [m["key"] for m in HRD_MENUS]
HRD_ACTIONS = ["view", "create", "edit", "delete", "report"]


def _now():
    return datetime.now(timezone.utc).isoformat()


def _is_super(current: dict) -> bool:
    return is_super_admin_user(current) or current.get("role") == "super_admin" or bool(current.get("is_super_admin"))


def has_perm(current: dict, menu: str, action: str) -> bool:
    acc = (current.get("access") or {}).get(menu) or {}
    if menu in GAJI_GROUP:
        # Area Gaji: akses eksplisit saja — Super Admin TIDAK otomatis punya akses.
        return bool(acc.get(action))
    if _is_super(current):
        return True
    return bool(acc.get(action))


def _has_any_hrd(current: dict) -> bool:
    if _is_super(current):
        return True
    acc = current.get("access") or {}
    return any((acc.get(k) or {}).get("view") for k in HRD_MENU_KEYS)


# Menu grup "gaji" — area sensitif yang dikunci PIN Gaji
GAJI_GROUP = {m["key"] for m in HRD_MENUS if m["group"] == "gaji"}


def _allowed_groups(current: dict) -> list:
    """Grup payroll yang boleh diakses user. Bos punya banyak grup."""
    g = current.get("payroll_groups")
    if isinstance(g, list) and g:
        return list(g)
    return [current.get("payroll_group") or "karyawan"]


def _is_boss(current: dict) -> bool:
    """Bos = user dengan akses lebih dari satu grup payroll (tanpa PIN Gaji)."""
    return len(_allowed_groups(current)) > 1


def _pgroup(current: dict) -> str:
    """Grup payroll aktif. Bos multi-grup memilih via header x-payroll-group.
    User biasa: 'karyawan' (Herliana, default) atau 'staff' (Nofia)."""
    allowed = _allowed_groups(current)
    ag = current.get("_active_group")
    if ag and ag in allowed:
        return ag
    return allowed[0]


def _pin_id(group: str) -> str:
    return "hrd" if group == "karyawan" else f"hrd_{group}"


def _pfilter(group: str) -> dict:
    """Filter isolasi slip per grup. 'karyawan' juga mencakup slip lama tanpa penanda."""
    if group == "karyawan":
        return {"$or": [{"payroll_group": "karyawan"}, {"payroll_group": {"$exists": False}}, {"payroll_group": None}]}
    return {"payroll_group": group}


def _can_manage_pin(current: dict) -> bool:
    """Hanya user yang punya akses gaji (Herliana/Nofia) yang boleh set/buat PIN Gaji.
    Super Admin TIDAK bisa membuat PIN Gaji — ia hanya menyetujui reset.
    Bos multi-grup TIDAK memakai PIN sama sekali."""
    if _is_boss(current):
        return False
    acc = current.get("access") or {}
    return any((acc.get(k) or {}).get("view") for k in GAJI_GROUP)


async def _gaji_pin_is_set(group: str = "karyawan") -> bool:
    s = await db.hrd_settings.find_one({"_id": _pin_id(group)})
    return bool(s and s.get("pin_hash"))


def _valid_token(token: str, scope: str, uid: str | None = None, group: str | None = None) -> bool:
    if not token:
        return False
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("scope") != scope:
            return False
        if uid is not None and payload.get("uid") != uid:
            return False
        if group is not None and (payload.get("group") or "karyawan") != group:
            return False
        return True
    except Exception:
        return False


# ---------------- Portal gate (akses HRD saja — TANPA PIN Portal) ----------------
async def require_hrd(x_payroll_group: str = Header(None), current: dict = Depends(get_current_user)) -> dict:
    if not _has_any_hrd(current):
        raise HTTPException(status_code=403, detail="Anda tidak punya akses ke Portal HRD")
    allowed = _allowed_groups(current)
    current["_active_group"] = x_payroll_group if (x_payroll_group and x_payroll_group in allowed) else allowed[0]
    return current


def require_hrd_perm(menu: str, action: str):
    """Permission menu/action. Untuk menu grup 'gaji', bila PIN Gaji aktif maka
    wajib token PIN Gaji (header x-hrd-gaji) sesuai GRUP payroll user."""
    async def _dep(x_hrd_gaji: str = Header(None), current: dict = Depends(require_hrd)) -> dict:
        if not has_perm(current, menu, action):
            raise HTTPException(status_code=403, detail="Anda tidak memiliki hak akses untuk aksi ini")
        if menu in GAJI_GROUP and not _is_boss(current):
            grp = _pgroup(current)
            if await _gaji_pin_is_set(grp) and not _valid_token(x_hrd_gaji, "hrd_gaji", group=grp):
                raise HTTPException(status_code=401, detail="PIN Gaji diperlukan")
        return current
    return _dep


class PinIn(BaseModel):
    pin: str
    current_pin: str | None = None


# ---------------- PIN Portal (per-user) ----------------
@router.post("/portal-pin/set")
async def set_portal_pin(payload: PinIn, current: dict = Depends(get_current_user)):
    if not _has_any_hrd(current):
        raise HTTPException(status_code=403, detail="Anda tidak punya akses ke Portal HRD")
    if not payload.pin or len(payload.pin) < 4:
        raise HTTPException(status_code=400, detail="PIN minimal 4 digit")
    existing = current.get("hrd_pin_hash")
    if existing and not (payload.current_pin and verify_password(payload.current_pin, existing)):
        raise HTTPException(status_code=400, detail="PIN portal lama salah")
    await db.users.update_one({"id": current["id"]}, {"$set": {"hrd_pin_hash": hash_password(payload.pin), "hrd_pin_updated_at": _now()}})
    await log_action(current, "hrd_set_portal_pin", "user", current["id"], {})
    return {"success": True}


@router.post("/portal-pin/verify")
async def verify_portal_pin(payload: PinIn, current: dict = Depends(get_current_user)):
    if not current.get("hrd_pin_hash"):
        raise HTTPException(status_code=400, detail="PIN Portal belum diatur.")
    if not verify_password(payload.pin, current["hrd_pin_hash"]):
        await log_action(current, "hrd_access_denied", "hrd", "", {"reason": "PIN portal salah"})
        raise HTTPException(status_code=401, detail="PIN Portal salah")
    token = jwt.encode({"scope": "hrd_portal", "uid": current.get("id"),
                        "exp": datetime.now(timezone.utc) + timedelta(hours=10)}, JWT_SECRET, algorithm=JWT_ALGORITHM)
    await log_action(current, "hrd_access", "hrd", "", {"portal": "HRD"})
    return {"portal_token": token}


# ---------------- PIN Gaji (khusus Herliana / gaji user) ----------------
@router.post("/set-pin")
async def set_pin(payload: PinIn, current: dict = Depends(get_current_user)):
    if not _can_manage_pin(current):
        raise HTTPException(status_code=403, detail="Hanya user Gaji (mis. Herliana) atau Super Admin yang bisa mengatur PIN Gaji")
    if not payload.pin or len(payload.pin) < 4:
        raise HTTPException(status_code=400, detail="PIN minimal 4 digit")
    grp = _pgroup(current)
    pid = _pin_id(grp)
    s = await db.hrd_settings.find_one({"_id": pid})
    # Bila ada persetujuan reset dari Super Admin (untuk grup ini), boleh set PIN baru TANPA PIN lama.
    approved = await db.hrd_pin_resets.count_documents({"status": "approved", "group": grp})
    if s and s.get("pin_hash") and not approved:
        if not (payload.current_pin and verify_password(payload.current_pin, s["pin_hash"])):
            raise HTTPException(status_code=400, detail="PIN Gaji lama salah")
    await db.hrd_settings.update_one({"_id": pid}, {"$set": {"pin_hash": hash_password(payload.pin), "pin_updated_at": _now()}}, upsert=True)
    # Bersihkan permintaan reset grup ini yang belum selesai
    await db.hrd_pin_resets.update_many({"group": grp, "status": {"$in": ["pending", "approved"]}}, {"$set": {"status": "resolved", "resolved_at": _now()}})
    await log_action(current, "hrd_set_pin", "hrd_settings", pid, {"via": "reset" if approved else "normal", "group": grp})
    return {"success": True}


@router.post("/verify-pin")
async def verify_pin(payload: PinIn, current: dict = Depends(require_hrd)):
    grp = _pgroup(current)
    s = await db.hrd_settings.find_one({"_id": _pin_id(grp)})
    if not s or not s.get("pin_hash"):
        raise HTTPException(status_code=400, detail="PIN Gaji belum diatur.")
    if not verify_password(payload.pin, s["pin_hash"]):
        await log_action(current, "hrd_access_denied", "hrd", "", {"reason": "PIN gaji salah"})
        raise HTTPException(status_code=401, detail="PIN Gaji salah")
    token = jwt.encode({"scope": "hrd_gaji", "group": grp, "uid": current.get("id"),
                        "exp": datetime.now(timezone.utc) + timedelta(hours=8)}, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return {"gaji_token": token}


@router.get("/pin-status")
async def pin_status(current: dict = Depends(get_current_user)):
    return {
        "portal_pin_set": bool(current.get("hrd_pin_hash")),
        "gaji_pin_set": await _gaji_pin_is_set(_pgroup(current)),
        "can_manage_gaji_pin": _can_manage_pin(current),
        "is_super": _is_super(current),
    }


@router.get("/menu-defs")
async def menu_defs(current: dict = Depends(get_current_user)):
    """Definisi menu HRD + aksi, untuk editor permission di panel Admin (Super Admin only)."""
    if not _is_super(current):
        raise HTTPException(status_code=403, detail="Hanya Super Admin")
    return {"menus": HRD_MENUS, "actions": HRD_ACTIONS}


@router.get("/my-access")
async def my_access(current: dict = Depends(get_current_user)):
    """Info untuk frontend: akses portal, status PIN (portal & gaji), dan matrix akses efektif."""
    is_super = _is_super(current)
    acc = current.get("access") or {}
    effective = {}
    for k in HRD_MENU_KEYS:
        if is_super and k not in GAJI_GROUP:
            effective[k] = {a: True for a in HRD_ACTIONS}
        else:
            m = acc.get(k) or {}
            effective[k] = {a: bool(m.get(a)) for a in HRD_ACTIONS}
    return {
        "is_super": is_super,
        "can_enter": is_super or _has_any_hrd(current),
        "gaji_pin_set": await _gaji_pin_is_set(_pgroup(current)),
        "can_manage_gaji_pin": _can_manage_pin(current),
        "can_approve_reset": is_super,
        "gaji_reset_pending": await db.hrd_pin_resets.count_documents({"status": "pending"}),
        "gaji_reset_approved": (await db.hrd_pin_resets.count_documents({"status": "approved", "group": _pgroup(current)})) > 0,
        "menus": HRD_MENUS,
        "gaji_group": sorted(GAJI_GROUP),
        "payroll_groups": _allowed_groups(current),
        "is_boss": _is_boss(current),
        "access": effective,
    }


# ---------------- PIN Gaji reset: Herliana ajukan → Susanto ACC → Herliana buat PIN baru ----------------
class ResetReqIn(BaseModel):
    reason: str | None = None


@router.post("/gaji-pin/request-reset")
async def request_gaji_pin_reset(payload: ResetReqIn, current: dict = Depends(require_hrd)):
    if not _can_manage_pin(current):
        raise HTTPException(status_code=403, detail="Hanya user Gaji (Herliana) yang bisa mengajukan reset PIN")
    grp = _pgroup(current)
    existing = await db.hrd_pin_resets.find_one({"group": grp, "status": {"$in": ["pending", "approved"]}})
    if existing:
        raise HTTPException(status_code=400, detail="Sudah ada permintaan reset yang sedang diproses")
    doc = {
        "id": str(uuid.uuid4()),
        "group": grp,
        "requested_by": current.get("id"),
        "requested_by_name": current.get("name") or current.get("username"),
        "reason": (payload.reason or "").strip(),
        "status": "pending",
        "created_at": _now(),
        "approved_at": None,
        "approved_by": None,
        "resolved_at": None,
    }
    await db.hrd_pin_resets.insert_one(dict(doc))
    await log_action(current, "hrd_pin_reset_request", "hrd", "", {"reason": doc["reason"]})
    return {"success": True, "request_id": doc["id"]}


@router.get("/gaji-pin/reset-requests")
async def list_gaji_pin_resets(current: dict = Depends(require_hrd)):
    if _is_super(current):
        items = await db.hrd_pin_resets.find({"status": "pending"}, {"_id": 0}).sort("created_at", -1).to_list(100)
    elif _can_manage_pin(current):
        items = await db.hrd_pin_resets.find({"status": {"$in": ["pending", "approved"]}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    else:
        raise HTTPException(status_code=403, detail="Tidak berwenang")
    return {"items": items, "count": len(items)}


class ResetActionIn(BaseModel):
    request_id: str | None = None


@router.post("/gaji-pin/approve-reset")
async def approve_gaji_pin_reset(payload: ResetActionIn, current: dict = Depends(get_current_user)):
    if not _is_super(current):
        raise HTTPException(status_code=403, detail="Hanya Super Admin (Susanto) yang bisa menyetujui reset PIN")
    flt = {"status": "pending"}
    if payload.request_id:
        flt["id"] = payload.request_id
    r = await db.hrd_pin_resets.update_many(flt, {"$set": {"status": "approved", "approved_at": _now(), "approved_by": current.get("id")}})
    if not r.modified_count:
        raise HTTPException(status_code=400, detail="Tidak ada permintaan reset yang menunggu")
    await log_action(current, "hrd_pin_reset_approve", "hrd", "", {"count": r.modified_count})
    return {"success": True, "approved": r.modified_count}


class ResetApplyIn(BaseModel):
    pin: str


@router.post("/gaji-pin/reset-apply")
async def apply_gaji_pin_reset(payload: ResetApplyIn, current: dict = Depends(require_hrd)):
    if not _can_manage_pin(current):
        raise HTTPException(status_code=403, detail="Hanya user Gaji (Herliana) yang bisa membuat PIN baru")
    grp = _pgroup(current)
    approved = await db.hrd_pin_resets.count_documents({"status": "approved", "group": grp})
    if not approved:
        raise HTTPException(status_code=400, detail="Belum ada persetujuan reset dari Super Admin (Susanto)")
    if not payload.pin or len(payload.pin) < 4:
        raise HTTPException(status_code=400, detail="PIN minimal 4 digit")
    await db.hrd_settings.update_one({"_id": _pin_id(grp)}, {"$set": {"pin_hash": hash_password(payload.pin), "pin_updated_at": _now()}}, upsert=True)
    await db.hrd_pin_resets.update_many({"status": "approved", "group": grp}, {"$set": {"status": "resolved", "resolved_at": _now()}})
    await log_action(current, "hrd_set_pin", "hrd_settings", _pin_id(grp), {"via": "reset", "group": grp})
    return {"success": True}


# ---------------- Access log ----------------
HRD_LOG_ACTIONS = ["hrd_access", "hrd_access_denied", "hrd_set_pin", "hrd_set_portal_pin", "hrd_import_excel", "hrd_blast", "hrd_pin_reset_request", "hrd_pin_reset_approve"]
ACTION_LABEL = {
    "hrd_access": "Buka Portal HRD",
    "hrd_access_denied": "Gagal masuk (PIN salah)",
    "hrd_set_pin": "Ubah/Set PIN Gaji",
    "hrd_set_portal_pin": "Ubah/Set PIN Portal",
    "hrd_import_excel": "Import Excel slip gaji",
    "hrd_blast": "Kirim email slip gaji",
    "hrd_pin_reset_request": "Ajukan Reset PIN Gaji",
    "hrd_pin_reset_approve": "Setujui Reset PIN Gaji",
}


@router.get("/logs")
async def hrd_logs(current: dict = Depends(require_hrd)):
    # Hanya pemegang wewenang PIN Gaji (Herliana). Heri & super admin tidak boleh.
    if not _can_manage_pin(current):
        raise HTTPException(status_code=403, detail="Anda tidak punya akses Log HRD")
    items = await db.activity_logs.find(
        {"action": {"$in": HRD_LOG_ACTIONS}}, {"_id": 0}
    ).sort("timestamp", -1).to_list(300)
    for it in items:
        it["action_label"] = ACTION_LABEL.get(it.get("action"), it.get("action"))
    return {"items": items}


# ---------------- Settings (SMTP fleksibel) ----------------
class SettingsIn(BaseModel):
    smtp_host: str | None = None
    smtp_port: int | None = None
    smtp_security: str | None = None  # "ssl" | "tls"
    smtp_username: str | None = None
    smtp_password: str | None = None
    sender_name: str | None = None
    email_subject: str | None = None
    email_body: str | None = None


@router.get("/settings")
async def get_settings(current: dict = Depends(require_hrd_perm("hrd_settings", "view"))):
    s = await db.hrd_settings.find_one({"_id": _pin_id(_pgroup(current))}) or {}
    return {
        "smtp_host": s.get("smtp_host", ""),
        "smtp_port": s.get("smtp_port", 465),
        "smtp_security": s.get("smtp_security", "ssl"),
        "smtp_username": s.get("smtp_username", ""),
        "has_smtp_password": bool(s.get("smtp_password")),
        "sender_name": s.get("sender_name", "PT. MITRA KARYA SARANA"),
        "email_subject": s.get("email_subject") or DEFAULT_EMAIL_SUBJECT,
        "email_body": s.get("email_body") or DEFAULT_EMAIL_BODY,
    }


@router.post("/settings")
async def save_settings(payload: SettingsIn, current: dict = Depends(require_hrd_perm("hrd_settings", "edit"))):
    upd = {}
    if payload.smtp_host is not None:
        upd["smtp_host"] = payload.smtp_host.strip()
    if payload.smtp_port is not None:
        upd["smtp_port"] = int(payload.smtp_port)
    if payload.smtp_security is not None:
        upd["smtp_security"] = "ssl" if payload.smtp_security.lower() == "ssl" else "tls"
    if payload.smtp_username is not None:
        upd["smtp_username"] = payload.smtp_username.strip()
    if payload.sender_name is not None:
        upd["sender_name"] = payload.sender_name.strip()
    if payload.email_subject is not None:
        upd["email_subject"] = payload.email_subject
    if payload.email_body is not None:
        upd["email_body"] = payload.email_body
    if payload.smtp_password:  # only overwrite when provided
        upd["smtp_password"] = payload.smtp_password.strip()
    upd["settings_updated_at"] = _now()
    await db.hrd_settings.update_one({"_id": _pin_id(_pgroup(current))}, {"$set": upd}, upsert=True)
    return {"success": True}


@router.post("/settings/test")
async def test_settings(current: dict = Depends(require_hrd_perm("hrd_settings", "edit"))):
    """Uji koneksi & login SMTP tanpa mengirim email."""
    s = await db.hrd_settings.find_one({"_id": _pin_id(_pgroup(current))}) or {}
    host, port = s.get("smtp_host"), int(s.get("smtp_port") or 465)
    security = s.get("smtp_security") or "ssl"
    user, pw = s.get("smtp_username"), s.get("smtp_password")
    if not host or not user or not pw:
        raise HTTPException(status_code=400, detail="Pengaturan SMTP belum lengkap. Isi Host, Username, dan Password.")
    try:
        server = _open_smtp(host, port, security, user, pw)
        try:
            server.quit()
        except Exception:
            pass
        return {"success": True, "message": f"Koneksi ke {host}:{port} berhasil. Login SMTP OK."}
    except Exception as e:
        raise HTTPException(status_code=400, detail=_smtp_friendly(str(e)))


# ---------------- Employees ----------------
class EmployeeIn(BaseModel):
    nik: str = ""
    nama: str = ""
    email: str = ""
    jabatan: str = ""
    no_rekening: str = ""
    bank: str = ""


@router.get("/employees")
async def list_employees(q: str = "", current: dict = Depends(require_hrd_perm("hrd_karyawan", "view"))):
    flt = dict(NOT_DELETED_FILTER)
    if q:
        flt["$and"] = [{"$or": [{"nama": {"$regex": q, "$options": "i"}}, {"nik": {"$regex": q, "$options": "i"}}, {"jabatan": {"$regex": q, "$options": "i"}}]}]
    items = await db.hrd_employees.find(flt, {"_id": 0}).sort("nama", 1).to_list(1000)
    return {"items": items}


@router.post("/employees")
async def create_employee(payload: EmployeeIn, current: dict = Depends(require_hrd_perm("hrd_karyawan", "create"))):
    if not payload.nama.strip():
        raise HTTPException(status_code=400, detail="Nama wajib diisi")
    doc = payload.dict()
    doc.update({"id": str(uuid.uuid4()), "active": True, "created_at": _now(), "updated_at": _now()})
    await db.hrd_employees.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


@router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, payload: EmployeeIn, current: dict = Depends(require_hrd_perm("hrd_karyawan", "edit"))):
    r = await db.hrd_employees.update_one({"id": emp_id, **NOT_DELETED_FILTER}, {"$set": {**payload.dict(), "updated_at": _now()}})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    return await db.hrd_employees.find_one({"id": emp_id}, {"_id": 0})


@router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, current: dict = Depends(require_hrd_perm("hrd_karyawan", "delete"))):
    ok = await soft_delete_one("hrd_employees", {"id": emp_id}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    return {"success": True}


# ---------------- Payslips ----------------
class Component(BaseModel):
    label: str = ""
    amount: float = 0.0
    qty: float | None = None
    unit: str | None = None


class PayslipIn(BaseModel):
    period_month: int
    period_year: int
    employee_id: str | None = None
    nik: str = ""
    nama: str = ""
    email: str = ""
    jabatan: str = ""
    dept: str = "Production"
    no_rekening: str = ""
    bank: str = ""
    earnings: list[Component] = []
    deductions: list[Component] = []
    take_home: float | None = None
    tanggal_lahir: str = ""
    notes: str = ""


def _round_rp(v: float) -> int:
    # Pembulatan ke ribuan terdekat
    return int(round(float(v or 0) / 1000.0) * 1000)


def _compute_slip(d: dict) -> dict:
    gross = round(sum(float(e.get("amount") or 0) for e in d.get("earnings", [])), 2)
    ded = round(sum(float(x.get("amount") or 0) for x in d.get("deductions", [])), 2)
    d["gross"] = gross
    d["total_deduction"] = ded
    d["net"] = round(gross - ded, 2)
    # take_home (PEMBULATAN) editable; default = pembulatan net ke ribuan
    if d.get("take_home") in (None, "", 0):
        d["take_home"] = _round_rp(d["net"])
    else:
        d["take_home"] = float(d["take_home"])
    return d


@router.get("/payslips")
async def list_payslips(month: int = 0, year: int = 0, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "view"))):
    flt = merged(NOT_DELETED_FILTER, _pfilter(_pgroup(current)))
    if month:
        flt["period_month"] = month
    if year:
        flt["period_year"] = year
    items = await db.hrd_payslips.find(flt, {"_id": 0}).sort("nama", 1).to_list(2000)
    return {"items": items}


@router.get("/payroll-summary")
async def payroll_summary(current: dict = Depends(require_hrd)):
    """Ringkasan take-home per grup untuk kartu Data Gaji bos (khusus bos multi-grup)."""
    if not _is_boss(current):
        raise HTTPException(status_code=403, detail="Hanya untuk akun bos multi-grup")
    out = []
    for g in _allowed_groups(current):
        flt = merged(NOT_DELETED_FILTER, _pfilter(g))
        docs = await db.hrd_payslips.find(flt, {"_id": 0, "take_home": 1, "period_month": 1, "period_year": 1}).to_list(5000)
        if not docs:
            out.append({"group": g, "count": 0, "total_take_home": 0, "period_month": 0, "period_year": 0})
            continue
        latest = max((d.get("period_year") or 0, d.get("period_month") or 0) for d in docs)
        sel = [d for d in docs if (d.get("period_year") or 0, d.get("period_month") or 0) == latest]
        total = sum(float(d.get("take_home") or 0) for d in sel)
        out.append({"group": g, "count": len(sel), "total_take_home": total, "period_month": latest[1], "period_year": latest[0]})
    return {"items": out}


@router.post("/payslips")
async def create_payslip(payload: PayslipIn, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "create"))):
    d = payload.dict()
    d["earnings"] = [e for e in d["earnings"] if (e.get("label") or e.get("amount"))]
    d["deductions"] = [e for e in d["deductions"] if (e.get("label") or e.get("amount"))]
    _compute_slip(d)
    d.update({"id": str(uuid.uuid4()), "email_status": "belum", "email_error": "", "sent_at": None, "created_at": _now(), "updated_at": _now()})
    await db.hrd_payslips.insert_one(dict(d))
    d.pop("_id", None)
    return d


@router.put("/payslips/{sid}")
async def update_payslip(sid: str, payload: PayslipIn, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "edit"))):
    d = payload.dict()
    d["earnings"] = [e for e in d["earnings"] if (e.get("label") or e.get("amount"))]
    d["deductions"] = [e for e in d["deductions"] if (e.get("label") or e.get("amount"))]
    _compute_slip(d)
    d["updated_at"] = _now()
    r = await db.hrd_payslips.update_one({"id": sid, **NOT_DELETED_FILTER}, {"$set": d})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    return await db.hrd_payslips.find_one({"id": sid}, {"_id": 0})


@router.delete("/payslips/{sid}")
async def delete_payslip(sid: str, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "delete"))):
    ok = await soft_delete_one("hrd_payslips", {"id": sid}, current)
    if not ok:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    return {"success": True}


DEDUCT_KEYWORDS = ["potong", "bpjs", "pph", "iuran", "pinjam", "kasbon", "deduct", "denda",
                   "koperasi", "absent", "absen", "jht", "jp", "jkk", "jkm", "jpk"]
KNOWN = {"nik": "nik", "kode": "nik", "kode_karyawan": "nik", "nama": "nama", "name": "nama",
         "email": "email", "jabatan": "jabatan", "posisi": "jabatan", "dept": "dept",
         "departemen": "dept", "department": "dept",
         "no_rekening": "no_rekening", "norekening": "no_rekening", "rekening": "no_rekening",
         "no_rek": "no_rekening", "bank": "bank"}


def _cell(ws, coord):
    try:
        return ws[coord].value
    except Exception:
        return None


def _numify(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except Exception:
        return None


def _norm_date(v):
    """Normalkan berbagai format tanggal ke ISO 'YYYY-MM-DD'."""
    if v is None:
        return ""
    if not isinstance(v, str) and hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return ""
    import re
    m = re.search(r"\b(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})\b", s)  # yyyy-mm-dd
    if m:
        y, mo, d = m.groups()
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return ""
    m = re.search(r"\b(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})\b", s)  # dd-mm-yyyy
    if m:
        d, mo, y = m.groups()
        y = ("20" + y) if len(y) == 2 else y
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return ""
    return ""


def _parse_directory(wb):
    """Baca sheet direktori (mis. 'Daftar Gaji') yang punya header NAMA & EMAIL.
    Kembalikan peta email + tanggal lahir per NAMA dan per NIK."""
    from openpyxl.utils import get_column_letter  # noqa
    by_nama, by_nik = {}, {}
    for name in wb.sheetnames:
        ws = wb[name]
        header_row, cols = None, {}
        for r in range(1, 13):
            labels = {}
            for c in range(1, 45):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    labels[v.strip().lower()] = c
            if "nama" in labels and "email" in labels:
                header_row = r
                cols = {
                    "nama": labels.get("nama"),
                    "email": labels.get("email"),
                    "nik": labels.get("nik"),
                    "lahir": next((labels[k] for k in labels if "lahir" in k), None),
                    "take_home": next((labels[k] for k in labels if "take home" in k), None),
                }
                break
        if not header_row:
            continue
        for r in range(header_row + 1, min(ws.max_row, 1000) + 1):
            nama = ws.cell(r, cols["nama"]).value if cols["nama"] else None
            if not nama or not str(nama).strip():
                continue
            email = ws.cell(r, cols["email"]).value if cols["email"] else None
            tgl = ws.cell(r, cols["lahir"]).value if cols.get("lahir") else None
            nik = ws.cell(r, cols["nik"]).value if cols.get("nik") else None
            th = ws.cell(r, cols["take_home"]).value if cols.get("take_home") else None
            rec = {
                "email": str(email).strip() if email else "",
                "tanggal_lahir": _norm_date(tgl),
                "nik": str(nik).strip() if nik else "",
                "take_home": _numify(th),
            }
            by_nama[str(nama).strip().lower()] = rec
            if rec["nik"]:
                by_nik[rec["nik"].lower()] = rec
    return by_nama, by_nik


def _parse_slip_sheet(ws, month, year):
    """Parse satu sheet slip (format cetak MKS) menjadi dict slip.
    Layout tetap: header di A1-A3, SLIP GAJI di A5, info di baris 8-10,
    PENGHASILAN (kolom A/C/E) & PENGURANGAN (kolom G/I/J/K)."""
    nama = _cell(ws, "C8")
    if not nama or str(nama).strip() == "":
        return None
    nik = _cell(ws, "E8")
    slip = {
        "nama": str(nama).strip(),
        "nik": str(nik).strip() if nik else "",
        "dept": str(_cell(ws, "C9") or "").strip(),
        "jabatan": str(_cell(ws, "C10") or "").strip(),
        "perhari": _numify(_cell(ws, "J8")),
        "lembur_jam": _numify(_cell(ws, "J9")),
        "tkehadiran_rate": _numify(_cell(ws, "J10")),
        "earnings": [],
        "deductions": [],
    }
    max_row = min(ws.max_row, 40)
    # Earnings: kolom A=label, C/D=qty, E=amount. Berhenti di "JUMLAH".
    gross = None
    for row in range(13, max_row + 1):
        lbl = _cell(ws, f"A{row}")
        lbl_s = str(lbl).strip() if lbl else ""
        if not lbl_s:
            continue
        if lbl_s.upper().startswith("JUMLAH"):
            gross = _numify(_cell(ws, f"E{row}"))
            break
        qty = _numify(_cell(ws, f"C{row}"))
        if qty is None:
            qty = _numify(_cell(ws, f"D{row}"))
        amt = _numify(_cell(ws, f"E{row}"))
        slip["earnings"].append({"label": lbl_s, "qty": qty, "amount": amt or 0})
    # Deductions: kolom G=label, I=qty, J=unit, K=amount.
    # Line item hanya dikumpulkan SEBELUM baris "JUMLAH"; setelah itu hanya
    # ambil PENGHASILAN BERSIH & PEMBULATAN lalu berhenti (abaikan footer).
    total_ded = None
    net = None
    take_home = None
    after_jumlah = False
    for row in range(13, max_row + 1):
        lbl = _cell(ws, f"G{row}")
        lbl_s = str(lbl).strip() if lbl else ""
        if not lbl_s:
            continue
        up = lbl_s.upper()
        if up.startswith("JUMLAH"):
            total_ded = _numify(_cell(ws, f"K{row}"))
            after_jumlah = True
            continue
        if "BERSIH" in up:
            net = _numify(_cell(ws, f"K{row}"))
            continue
        if "PEMBULATAN" in up:
            take_home = _numify(_cell(ws, f"K{row}"))
            break
        if after_jumlah:
            continue  # abaikan footer (Batam, Prepared By, HRD, dll)
        qty = _numify(_cell(ws, f"I{row}"))
        unit = _cell(ws, f"J{row}")
        unit_s = str(unit).strip() if unit and not isinstance(unit, (int, float)) else ""
        amt = _numify(_cell(ws, f"K{row}"))
        if lbl_s.strip().lower() == "pinjaman":
            lbl_s = "Lain-lain"
        slip["deductions"].append({"label": lbl_s, "qty": qty, "unit": unit_s, "amount": amt or 0})
    # Terbilang (cari di kolom A/B/C sekitar baris 26)
    terbilang = None
    for row in range(24, max_row + 1):
        a = _cell(ws, f"A{row}")
        if a and "terbilang" in str(a).lower():
            terbilang = _cell(ws, f"C{row}") or _cell(ws, f"D{row}")
            break
    slip["gross"] = gross if gross is not None else round(sum((e.get("amount") or 0) for e in slip["earnings"]), 2)
    slip["total_deduction"] = total_ded if total_ded is not None else round(sum((d.get("amount") or 0) for d in slip["deductions"]), 2)
    slip["net"] = net if net is not None else round(slip["gross"] - slip["total_deduction"], 2)
    slip["take_home"] = take_home if take_home is not None else _round_rp(slip["net"])
    slip["terbilang"] = str(terbilang).strip() if terbilang else ""
    # Deteksi email opsional: cari sel berisi alamat email di mana pun pada sheet slip
    email_sheet = ""
    for r in range(6, max_row + 1):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "@" in v:
                cand = v.strip()
                local, _, dom = cand.partition("@")
                if local and "." in dom and " " not in cand:
                    email_sheet = cand
                    break
        if email_sheet:
            break
    slip["email_sheet"] = email_sheet
    # Deteksi tanggal lahir: cari label mengandung 'lahir' lalu ambil nilai tanggal di kanannya
    tgl_lahir = ""
    for r in range(6, max_row + 1):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "lahir" in v.lower():
                nd = _norm_date(v)
                if nd:
                    tgl_lahir = nd
                else:
                    for cc in range(c + 1, 14):
                        nd = _norm_date(ws.cell(r, cc).value)
                        if nd:
                            tgl_lahir = nd
                            break
            if tgl_lahir:
                break
        if tgl_lahir:
            break
    slip["tanggal_lahir"] = tgl_lahir
    slip["period_month"] = month
    slip["period_year"] = year
    return slip


@router.post("/payslips/import-excel")
async def import_excel(month: int = Form(...), year: int = Form(...), file: UploadFile = File(...), current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "create"))):
    from openpyxl import load_workbook
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File Excel tidak valid (.xlsx)")

    # Deteksi sheet slip: cell A5 == "SLIP GAJI"
    slip_sheets = [n for n in wb.sheetnames if str(_cell(wb[n], "A5") or "").strip().upper() == "SLIP GAJI"]
    if not slip_sheets:
        raise HTTPException(status_code=400,
                            detail="Tidak ditemukan sheet slip gaji (yang memiliki judul 'SLIP GAJI'). Pastikan file berisi sheet slip per karyawan.")

    # Preload master karyawan untuk auto-match email/bank/rekening
    emps = await db.hrd_employees.find(NOT_DELETED_FILTER, {"_id": 0}).to_list(2000)
    by_nik = {(e.get("nik") or "").strip().lower(): e for e in emps if e.get("nik")}
    by_nama = {(e.get("nama") or "").strip().lower(): e for e in emps if e.get("nama")}
    # Baca tabel direktori (sheet 'Daftar Gaji') untuk email + tanggal lahir
    dir_by_nama, dir_by_nik = _parse_directory(wb)

    created = 0
    updated = 0
    names = []
    for name in slip_sheets:
        slip = _parse_slip_sheet(wb[name], month, year)
        if not slip:
            continue
        # Auto-match ke Master Karyawan (email/bank/rekening)
        match = by_nik.get((slip.get("nik") or "").strip().lower()) if slip.get("nik") else None
        if not match:
            match = by_nama.get(slip["nama"].strip().lower())
        sheet_email = (slip.pop("email_sheet", "") or "").strip()
        tgl_sheet = (slip.get("tanggal_lahir") or "").strip()  # dari sheet slip (jika ada)
        slip["email"] = ""
        slip["bank"] = ""
        slip["no_rekening"] = ""
        if match:
            slip["employee_id"] = match.get("id")
            slip["email"] = match.get("email", "")
            slip["bank"] = match.get("bank", "")
            slip["no_rekening"] = match.get("no_rekening", "")
            if not slip.get("jabatan") and match.get("jabatan"):
                slip["jabatan"] = match["jabatan"]
        # Tabel direktori 'Daftar Gaji' (cocokkan by NIK lalu by NAMA)
        drec = dir_by_nik.get((slip.get("nik") or "").strip().lower()) or dir_by_nama.get(slip["nama"].strip().lower())
        if drec and drec.get("email"):
            slip["email"] = drec["email"]
        # Tanggal lahir: prioritas sheet slip > tabel direktori
        if tgl_sheet:
            slip["tanggal_lahir"] = tgl_sheet
        elif drec and drec.get("tanggal_lahir"):
            slip["tanggal_lahir"] = drec["tanggal_lahir"]
        # Email dari sheet slip paling spesifik (override)
        if sheet_email:
            slip["email"] = sheet_email
        # Audit: bandingkan Take Home slip vs tabel 'Daftar Gaji' (kolom Take Home Pay)
        dg_th = drec.get("take_home") if drec else None
        slip["dg_take_home"] = dg_th
        if dg_th is not None and slip.get("take_home") is not None:
            slip["audit_diff"] = round(float(slip["take_home"]) - float(dg_th), 2)
            slip["audit_mismatch"] = abs(slip["audit_diff"]) >= 1
        else:
            slip["audit_diff"] = None
            slip["audit_mismatch"] = False
        # Isolasi grup payroll (karyawan=Herliana, staff=Nofia)
        grp = _pgroup(current)
        slip["payroll_group"] = grp
        # Upsert berdasarkan (period + nama + grup) agar re-import menimpa, bukan dobel
        existing = await db.hrd_payslips.find_one(
            merged({"period_month": month, "period_year": year, "nama": slip["nama"]}, NOT_DELETED_FILTER, _pfilter(grp)))
        slip["updated_at"] = _now()
        if existing:
            slip["email_status"] = existing.get("email_status", "belum")
            slip["email_error"] = existing.get("email_error", "")
            # Pertahankan email yang sudah diisi manual bila hasil parse kosong
            if not slip.get("email") and existing.get("email"):
                slip["email"] = existing["email"]
            if not slip.get("tanggal_lahir") and existing.get("tanggal_lahir"):
                slip["tanggal_lahir"] = existing["tanggal_lahir"]
            await db.hrd_payslips.update_one({"id": existing["id"]}, {"$set": slip})
            updated += 1
        else:
            slip.update({"id": str(uuid.uuid4()), "email_status": "belum", "email_error": "",
                         "sent_at": None, "notes": "", "created_at": _now()})
            await db.hrd_payslips.insert_one(dict(slip))
            created += 1
        names.append(slip["nama"])

    await log_action(current, "hrd_import_excel", "hrd_payslips", "",
                     {"created": created, "updated": updated, "period": f"{month}-{year}", "sheets": len(slip_sheets)})
    return {"success": True, "created": created, "updated": updated, "names": names, "sheets": slip_sheets}


# ---------------- PDF slip ----------------


def _rp(v):
    """Angka Rupiah format Indonesia dengan 2 desimal; 0/None => '-' (untuk isi teks biasa)."""
    try:
        f = float(v or 0)
    except Exception:
        return "Rp -"
    if abs(f) < 0.005:
        return "Rp -"
    s = f"{f:,.2f}".replace(",", "#").replace(".", ",").replace("#", ".")
    return "Rp " + s


def _money(v):
    return _rp(v)


def _qty(v):
    if v is None or v == "":
        return ""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else f"{f:g}"
    except Exception:
        return str(v)


_SATUAN = ["", "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh", "delapan", "sembilan",
           "sepuluh", "sebelas"]


def _terbilang(n: int) -> str:
    n = int(abs(n))
    if n < 12:
        return _SATUAN[n]
    if n < 20:
        return _terbilang(n - 10) + " belas"
    if n < 100:
        return _terbilang(n // 10) + " puluh" + ((" " + _terbilang(n % 10)) if n % 10 else "")
    if n < 200:
        return "seratus" + ((" " + _terbilang(n - 100)) if n - 100 else "")
    if n < 1000:
        return _terbilang(n // 100) + " ratus" + ((" " + _terbilang(n % 100)) if n % 100 else "")
    if n < 2000:
        return "seribu" + ((" " + _terbilang(n - 1000)) if n - 1000 else "")
    if n < 1_000_000:
        return _terbilang(n // 1000) + " ribu" + ((" " + _terbilang(n % 1000)) if n % 1000 else "")
    if n < 1_000_000_000:
        return _terbilang(n // 1_000_000) + " juta" + ((" " + _terbilang(n % 1_000_000)) if n % 1_000_000 else "")
    return _terbilang(n // 1_000_000_000) + " milyar" + ((" " + _terbilang(n % 1_000_000_000)) if n % 1_000_000_000 else "")


def _terbilang_rupiah(n) -> str:
    words = _terbilang(int(round(float(n or 0)))).strip()
    if not words:
        words = "nol"
    return " ".join(w.capitalize() for w in words.split()) + " Rupiah"


def _slip_no_dok(slip: dict) -> str:
    return f"SG/{slip.get('period_year')}/{int(slip.get('period_month') or 0):02d}/{(slip.get('nik') or '-').replace(' ', '')}"


def _slip_kode(slip: dict) -> str:
    """Kode verifikasi HMAC dari id + periode + take_home (stabil per slip)."""
    base = f"{slip.get('id')}|{slip.get('period_month')}-{slip.get('period_year')}|{int(round(float(slip.get('take_home') or 0)))}"
    h = hmac.new(JWT_SECRET.encode(), base.encode(), hashlib.sha256).hexdigest()[:12].upper()
    return f"{h[:4]}-{h[4:8]}-{h[8:]}"


def _slip_qr(slip: dict, no_dok: str, kode: str):
    import qrcode
    base = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
    content = f"{base}/verify?kode={kode}" if base else f"Slip Gaji MKS | {no_dok} | Kode: {kode}"
    img = qrcode.make(content, box_size=8, border=2)
    b = io.BytesIO()
    img.save(b, format="PNG")
    b.seek(0)
    return b


def _watermark(canvas, doc):
    """Footer 1 baris (tanpa watermark diagonal)."""
    from reportlab.lib.pagesizes import A4
    canvas.saveState()
    w2, h2 = A4
    canvas.setStrokeColorRGB(0.80, 0.84, 0.88)
    canvas.setLineWidth(0.5)
    x0 = 16 * 2.8346
    canvas.line(x0, 34, w2 - x0, 34)
    footer = ("Dokumen ini diterbitkan secara otomatis melalui HRIS PT Mitra Karya Sarana dan sah tanpa memerlukan tanda tangan basah. "
              "| RAHASIA — Dokumen pribadi, tidak untuk disebarluaskan.")
    avail = (w2 - 2 * x0) - 42  # sisakan ruang utk "Hal. N" di kanan
    fs = 6.8
    while fs > 4.5 and canvas.stringWidth(footer, "Helvetica", fs) > avail:
        fs -= 0.1
    canvas.setFont("Helvetica", fs)
    canvas.setFillColorRGB(0.42, 0.45, 0.50)
    canvas.drawString(x0, 24, footer)
    canvas.setFont("Helvetica", 6.8)
    canvas.drawRightString(w2 - x0, 24, f"Hal. {canvas.getPageNumber()}")
    canvas.restoreState()


def _render_slip_pdf(slip: dict, printed_by: str = "") -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=14 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=9)
    tiny = ParagraphStyle("tn", parent=styles["Normal"], fontSize=7.5, textColor=colors.HexColor("#475569"))
    elems = []
    GREY = colors.HexColor("#334155")
    LINE = colors.HexColor("#94A3B8")
    DARK = colors.HexColor("#1E293B")
    LIGHT = colors.HexColor("#F1F5F9")
    ORANGE_BG = colors.HexColor("#FFF7ED")
    ORANGE = colors.HexColor("#C2410C")
    CW = 178 * mm  # lebar konten total

    # Header teks polos (tanpa logo/kop — aman dari penyalahgunaan)
    head = Table([
        [Paragraph("PT. MITRA KARYA SARANA",
                   ParagraphStyle("hc", parent=styles["Normal"], fontSize=12, alignment=1,
                                  fontName="Helvetica-Bold", textColor=DARK))],
        [Paragraph("Taiwan International Park Blok B No. 117 - Kel. Kabil, Kec. Nongsa, Kota Batam, Kepulauan Riau",
                   ParagraphStyle("hs", parent=styles["Normal"], fontSize=8, alignment=1,
                                  textColor=colors.HexColor("#64748B")))],
    ], colWidths=[CW])
    head.setStyle(TableStyle([
        ("LINEBELOW", (0, 1), (0, 1), 0.8, DARK),
        ("BOTTOMPADDING", (0, 0), (0, 0), 1), ("BOTTOMPADDING", (0, 1), (0, 1), 5),
    ]))
    elems.append(head)
    elems.append(Spacer(1, 8))

    # Judul: band gelap "SLIP GAJI" + baris periode
    per = f"{BULAN_ID.get(slip.get('period_month'), slip.get('period_month'))} {slip.get('period_year')}"
    title = Table([
        [Paragraph("SLIP GAJI", ParagraphStyle("t", parent=styles["Normal"], fontSize=13,
                                               alignment=1, fontName="Helvetica-Bold",
                                               textColor=colors.white))],
        [Paragraph(f"Periode : <b>{per}</b>", ParagraphStyle("per", parent=small, alignment=1, fontSize=9.5))],
    ], colWidths=[CW])
    title.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), DARK),
        ("BACKGROUND", (0, 1), (0, 1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.8, DARK),
        ("TOPPADDING", (0, 0), (0, 0), 5), ("BOTTOMPADDING", (0, 0), (0, 0), 5),
        ("TOPPADDING", (0, 1), (0, 1), 3), ("BOTTOMPADDING", (0, 1), (0, 1), 3),
    ]))
    elems.append(title)
    elems.append(Spacer(1, 8))

    # Info karyawan (kiri) + rate (kanan)
    nik = slip.get("nik", "")
    info = Table([
        ["Nama", ":", slip.get("nama", ""), "Perhari", ":", _money(slip.get("perhari"))],
        ["NIK", ":", nik, "Lembur/Jam", ":", _money(slip.get("lembur_jam"))],
        ["Dept", ":", slip.get("dept", "") or "Production", "T. Kehadiran", ":", _money(slip.get("tkehadiran_rate"))],
        ["Jabatan", ":", slip.get("jabatan", ""), "", "", ""],
    ], colWidths=[20 * mm, 4 * mm, 66 * mm, 28 * mm, 4 * mm, 56 * mm])
    info.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (2, 0), (2, 0), "Helvetica-Bold"),
        ("ALIGN", (5, 0), (5, -1), "RIGHT"),
        ("TEXTCOLOR", (0, 0), (0, -1), GREY), ("TEXTCOLOR", (3, 0), (3, -1), GREY),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5), ("TOPPADDING", (0, 0), (-1, -1), 1.5),
    ]))
    elems.append(info)
    elems.append(Spacer(1, 7))

    # Tabel PENGHASILAN | PENGURANGAN
    earns = slip.get("earnings", []) or []
    deds = slip.get("deductions", []) or []
    rows = [["PENGHASILAN", "", "", "PENGURANGAN", "", ""]]
    n = max(len(earns), len(deds))
    for i in range(n):
        e = earns[i] if i < len(earns) else None
        d = deds[i] if i < len(deds) else None
        d_qty = ""
        if d is not None:
            q = _qty(d.get("qty"))
            u = (d.get("unit") or "").strip()
            d_qty = (q + (f" {u}" if u else "")).strip()
        rows.append([
            (e.get("label", "") if e else ""), (_qty(e.get("qty")) if e else ""), (_money(e.get("amount")) if e else ""),
            (d.get("label", "") if d else ""), d_qty, (_money(d.get("amount")) if d else ""),
        ])
    rows.append(["JUMLAH", "", _money(slip.get("gross")), "JUMLAH", "", _money(slip.get("total_deduction"))])
    rows.append(["", "", "", "PENGHASILAN BERSIH", "", _money(slip.get("net"))])
    rows.append(["", "", "", "TAKE HOME PAY", "", _money(slip.get("take_home"))])

    jml_row = len(rows) - 3
    t = Table(rows, colWidths=[36 * mm, 12 * mm, 41 * mm, 34 * mm, 14 * mm, 41 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        # header
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), LIGHT),
        ("TEXTCOLOR", (0, 0), (-1, 0), DARK),
        ("TOPPADDING", (0, 0), (-1, 0), 4), ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        # alignment
        ("ALIGN", (2, 0), (2, -1), "RIGHT"), ("ALIGN", (5, 0), (5, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"), ("ALIGN", (4, 0), (4, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # garis
        ("LINEBELOW", (0, 0), (-1, 0), 0.7, GREY),
        ("LINEBELOW", (0, 1), (-1, jml_row - 1), 0.25, colors.HexColor("#E2E8F0")),
        ("LINEABOVE", (0, jml_row), (-1, jml_row), 0.7, GREY),
        # JUMLAH
        ("FONTNAME", (0, jml_row), (-1, jml_row), "Helvetica-Bold"),
        ("BACKGROUND", (0, jml_row), (-1, jml_row), LIGHT),
        # PENGHASILAN BERSIH & TAKE HOME
        ("FONTNAME", (3, jml_row + 1), (5, jml_row + 2), "Helvetica-Bold"),
        ("LINEABOVE", (3, jml_row + 1), (5, jml_row + 1), 0.4, GREY),
        ("BACKGROUND", (3, jml_row + 2), (5, jml_row + 2), ORANGE_BG),
        ("TEXTCOLOR", (3, jml_row + 2), (5, jml_row + 2), ORANGE),
        ("FONTSIZE", (3, jml_row + 2), (5, jml_row + 2), 9.5),
        ("BOX", (3, jml_row + 2), (5, jml_row + 2), 0.8, ORANGE),
        # box kolom
        ("BOX", (0, 0), (2, jml_row), 0.8, GREY),
        ("BOX", (3, 0), (5, jml_row + 2), 0.8, GREY),
        ("LINEBEFORE", (3, 0), (3, jml_row), 0.8, GREY),
        ("TOPPADDING", (0, 1), (-1, -1), 3), ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 8))

    # Terbilang dalam kotak
    terb = slip.get("terbilang") or _terbilang_rupiah(slip.get("take_home"))
    terb_tbl = Table([[Paragraph(f"<b>Terbilang</b> : <i>{terb}</i>", small)]], colWidths=[CW])
    terb_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, LINE), ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elems.append(terb_tbl)

    if slip.get("notes"):
        elems.append(Spacer(1, 5))
        elems.append(Paragraph(f"Catatan : {slip['notes']}", small))

    # Blok validasi digital (tanpa QR — slip gaji tidak memakai QR sesuai kebijakan manajemen)
    elems.append(Spacer(1, 12))
    tgl = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7)))
    tgl_str = f"{tgl.day} {BULAN_ID.get(tgl.month)} {tgl.year}"
    stamp = tgl.strftime("%d-%m-%Y %H:%M") + " WIB"
    no_dok = _slip_no_dok(slip)
    kode = slip.get("kode") or _slip_kode(slip)
    note_style = ParagraphStyle("nv", parent=styles["Normal"], fontSize=7.5, textColor=GREY, leading=10)
    info_cell = Paragraph(
        "<b><font size=8>VALIDASI DOKUMEN ELEKTRONIK</font></b><br/>"
        f"Kode Verifikasi : <b>{kode}</b><br/>"
        f"Diterbitkan oleh HRD — PT. Mitra Karya Sarana<br/>"
        f"Batam, {tgl_str} · {stamp}",
        note_style)
    valid_tbl = Table([[info_cell]], colWidths=[CW])
    valid_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, LINE),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elems.append(valid_tbl)

    pdf.build(elems, onFirstPage=_watermark, onLaterPages=_watermark)
    buf.seek(0)
    return buf


@router.get("/payslips/{sid}/pdf")
async def payslip_pdf(sid: str, current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "report"))):
    slip = await db.hrd_payslips.find_one({"id": sid, **NOT_DELETED_FILTER}, {"_id": 0})
    if not slip:
        raise HTTPException(status_code=404, detail="Slip tidak ditemukan")
    if not slip.get("kode"):
        slip["kode"] = _slip_kode(slip)
        await db.hrd_payslips.update_one({"id": sid}, {"$set": {"kode": slip["kode"]}})
    buf = _render_slip_pdf(slip, printed_by=current.get("name") or current.get("username", ""))
    fname = f"SlipGaji_{slip.get('nama','')}_{slip.get('period_month')}_{slip.get('period_year')}.pdf".replace(" ", "_")
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{fname}"'})


class SlipVerifyIn(BaseModel):
    kode: str


@router.post("/payslips/verify")
async def verify_payslip(payload: SlipVerifyIn, current: dict = Depends(require_hrd_perm("hrd_dokumen", "view"))):
    raw = "".join(ch for ch in payload.kode.upper() if ch.isalnum())
    if len(raw) != 12:
        return {"valid": False, "message": "Format kode tidak valid (12 karakter)"}
    kode = f"{raw[:4]}-{raw[4:8]}-{raw[8:]}"
    slip = await db.hrd_payslips.find_one({"kode": kode, **NOT_DELETED_FILTER}, {"_id": 0})
    if not slip:
        return {"valid": False, "message": "Kode slip gaji tidak terdaftar / tidak sah"}
    per = f"{BULAN_ID.get(slip.get('period_month'), slip.get('period_month'))} {slip.get('period_year')}"
    return {"valid": True, "slip": {
        "no_dok": _slip_no_dok(slip), "nama": slip.get("nama", ""), "nik": slip.get("nik", ""),
        "periode": per, "take_home": slip.get("take_home"), "kode": kode}}


@router.get("/verify/{kode}")
async def public_verify(kode: str):
    """Verifikasi keaslian surat/slip secara PUBLIK (tanpa login) — untuk QR online."""
    raw = "".join(ch for ch in str(kode).upper() if ch.isalnum())
    if len(raw) != 12:
        return {"valid": False, "message": "Format kode tidak valid"}
    k = f"{raw[:4]}-{raw[4:8]}-{raw[8:]}"
    letter = await db.hrd_letters.find_one({"kode": k, **NOT_DELETED_FILTER}, {"_id": 0})
    if letter:
        from routers.hrd_people import LETTER_KINDS
        return {"valid": True, "type": "letter", "data": {
            "nomor": letter.get("nomor"), "jenis": LETTER_KINDS.get(letter.get("jenis"), {}).get("title", letter.get("jenis")),
            "nama": letter.get("nama", ""), "nik": letter.get("nik", ""), "jabatan": letter.get("jabatan", ""),
            "created_at": letter.get("created_at"), "created_by": letter.get("created_by", "")}}
    slip = await db.hrd_payslips.find_one({"kode": k, **NOT_DELETED_FILTER}, {"_id": 0})
    if slip:
        per = f"{BULAN_ID.get(slip.get('period_month'), slip.get('period_month'))} {slip.get('period_year')}"
        return {"valid": True, "type": "slip", "data": {
            "no_dok": _slip_no_dok(slip), "nama": slip.get("nama", ""), "nik": slip.get("nik", ""),
            "periode": per, "take_home": slip.get("take_home")}}
    return {"valid": False, "message": "Kode tidak terdaftar. Dokumen tidak dikenali sistem HRD."}


# Kolom penghasilan & potongan sesuai template slip gaji MKS
SLIP_EARNINGS = ["Gaji Pokok", "T. Tetap", "T. Kehadiran", "Lembur (1.5)", "Lembur (2)",
                 "Lembur (3)", "Lembur (4)", "Insentive+ 2nd Shift"]
SLIP_DEDUCTIONS = ["Absent", "T. Transport", "PPh 21", "JHT+JP (2%+1%)", "BPJS KESEHATAN 1%", "Lain-lain"]


@router.get("/import-template")
async def import_template(current: dict = Depends(require_hrd_perm("hrd_slip_gaji", "view"))):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Slip Gaji"
    headers = ["NIK", "Nama", "Email", "Jabatan", "Dept", "No Rekening", "Bank"] + SLIP_EARNINGS + SLIP_DEDUCTIONS
    ws.append(headers)
    ws.append(["MKS 0021", "Wawan Munandar", "wawan@email.com", "Supervisor", "Production", "1234567890", "BCA",
               5000000, 570000, 0, 0, 0, 0, 0, 0,  # earnings
               0, 0, 91634, 304920, 101640, 0])     # deductions
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="Template_Import_Slip_Gaji.xlsx"'})


@router.get("/slip-labels")
async def slip_labels(current: dict = Depends(require_hrd)):
    return {"earnings": SLIP_EARNINGS, "deductions": SLIP_DEDUCTIONS}


# ---------------- Blast email ----------------
class BlastIn(BaseModel):
    month: int
    year: int
    ids: list[str] | None = None


@router.post("/blast")
async def blast(payload: BlastIn, current: dict = Depends(require_hrd_perm("hrd_email", "create"))):
    grp = _pgroup(current)
    s = await db.hrd_settings.find_one({"_id": _pin_id(grp)}) or {}
    smtp_host = s.get("smtp_host")
    smtp_port = int(s.get("smtp_port") or 465)
    smtp_security = s.get("smtp_security") or "ssl"
    smtp_user = s.get("smtp_username")
    smtp_pw = s.get("smtp_password")
    sender_name = s.get("sender_name") or "PT. MITRA KARYA SARANA"
    subj_tpl = s.get("email_subject") or DEFAULT_EMAIL_SUBJECT
    body_tpl = s.get("email_body") or DEFAULT_EMAIL_BODY
    if not smtp_host or not smtp_user or not smtp_pw:
        raise HTTPException(status_code=400, detail="SMTP belum dikonfigurasi. Isi Host, Username & Password di tab Pengaturan Email.")

    flt = merged({"period_month": payload.month, "period_year": payload.year}, NOT_DELETED_FILTER, _pfilter(grp))
    if payload.ids:
        flt["id"] = {"$in": payload.ids}
    slips = await db.hrd_payslips.find(flt, {"_id": 0}).to_list(2000)
    if not slips:
        raise HTTPException(status_code=400, detail="Tidak ada slip untuk periode ini")

    per_label = f"{BULAN_ID.get(payload.month, payload.month)} {payload.year}"
    results = []
    try:
        server = _open_smtp(smtp_host, smtp_port, smtp_security, smtp_user, smtp_pw)
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=400, detail="Login SMTP gagal. Periksa Username & Password email.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=_smtp_friendly(str(e)))

    try:
        for slip in slips:
            email_to = (slip.get("email") or "").strip()
            status, err = "gagal", ""
            if not email_to:
                err = "Email kosong"
            else:
                try:
                    tvars = _SafeDict({
                        "nama": slip.get("nama", ""), "nik": slip.get("nik", ""),
                        "jabatan": slip.get("jabatan", ""), "dept": slip.get("dept", ""),
                        "bulan": BULAN_ID.get(payload.month, payload.month), "tahun": payload.year,
                        "periode": per_label, "take_home": _rp(slip.get("take_home")),
                        "sender": sender_name,
                    })
                    msg = MIMEMultipart()
                    msg["From"] = f"{sender_name} <{smtp_user}>"
                    msg["To"] = email_to
                    msg["Date"] = formatdate(localtime=True)
                    try:
                        msg["Message-ID"] = make_msgid(domain=smtp_user.split("@")[-1])
                    except Exception:
                        pass
                    msg["Subject"] = subj_tpl.format_map(tvars)
                    body = body_tpl.format_map(tvars)
                    msg.attach(MIMEText(body, "plain"))
                    pdf_buf = _render_slip_pdf(slip, printed_by=current.get("name") or current.get("username", ""))
                    part = MIMEApplication(pdf_buf.read(), _subtype="pdf")
                    part.add_header("Content-Disposition", "attachment",
                                    filename=f"SlipGaji_{per_label}_{slip.get('nama','')}.pdf".replace(" ", "_"))
                    msg.attach(part)
                    server.sendmail(smtp_user, [email_to], msg.as_string())
                    status, err = "terkirim", ""
                except Exception as e:
                    status, err = "gagal", _smtp_friendly(str(e))
            await db.hrd_payslips.update_one({"id": slip["id"]}, {"$set": {"email_status": status, "email_error": err, "sent_at": _now() if status == "terkirim" else None}})
            results.append({"id": slip["id"], "nama": slip.get("nama"), "email": email_to, "status": status, "error": err})
    finally:
        try:
            server.quit()
        except Exception:
            pass

    sent = sum(1 for r in results if r["status"] == "terkirim")
    await log_action(current, "hrd_blast", "hrd_payslips", "", {"period": per_label, "sent": sent, "total": len(results)})
    return {"success": True, "sent": sent, "failed": len(results) - sent, "results": results}
