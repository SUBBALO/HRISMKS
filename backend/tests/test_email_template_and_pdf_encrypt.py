"""Iteration 4: Editable email template + birth-date PDF encryption tests.

Covers:
- GET/POST /api/hrd/settings for email_subject/email_body (defaults + persistence)
- PUT /api/hrd/payslips/{id} tanggal_lahir persists
- Importer auto-detects tanggal_lahir from a slip sheet cell containing 'lahir'
- Re-import preserves manually-set tanggal_lahir when sheet has none
- _render_slip_pdf encrypts with birth-date password (contains b'/Encrypt') when password given
- GET /api/hrd/payslips/{id}/pdf preview is 200 application/pdf (unencrypted preview)
- POST /api/hrd/blast returns 400 clear error when Gmail App Password not configured
"""
import io
import os
import shutil
import asyncio
import pytest
import requests

BASE = (os.environ.get("REACT_APP_BACKEND_URL") or "").rstrip("/")
if not BASE:
    with open("/app/frontend/.env") as f:
        for ln in f:
            if ln.startswith("REACT_APP_BACKEND_URL="):
                BASE = ln.split("=", 1)[1].strip().rstrip("/")
API = f"{BASE}/api"
XLSX_SRC = "/tmp/gaji.xlsx"

HER = ("herliana", "123456")
GAJI_PIN = "5678"

DEFAULT_SUBJECT = "Slip Gaji {bulan} {tahun} - {nama}"


def _login(u, p):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": u, "password": p}, timeout=15)
    assert r.status_code == 200, f"login {u} -> {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def her_sess():
    s = _login(*HER)
    v = s.post(f"{API}/hrd/verify-pin", json={"pin": GAJI_PIN})
    assert v.status_code == 200, f"gaji verify: {v.status_code} {v.text}"
    s.headers.update({"x-hrd-gaji": v.json()["gaji_token"]})
    return s


# ---------- Email template settings ----------
def test_settings_default_template_present(her_sess):
    r = her_sess.get(f"{API}/hrd/settings")
    assert r.status_code == 200, r.text
    d = r.json()
    assert "email_subject" in d and "email_body" in d
    assert d["email_subject"].strip() != ""
    assert d["email_body"].strip() != ""
    # default subject template contains {nama}
    assert "{nama}" in d["email_subject"] or "{nama}" in d["email_body"]


def test_settings_save_and_reload_template(her_sess):
    new_subj = "TEST Slip {bulan} {tahun} - {nama}"
    new_body = "Halo {nama} ({nik}) — {jabatan}\nPeriode {bulan} {tahun}\nTake home: {take_home}\n\n{sender}"
    r = her_sess.post(f"{API}/hrd/settings", json={
        "email_subject": new_subj,
        "email_body": new_body,
    })
    assert r.status_code == 200, r.text
    g = her_sess.get(f"{API}/hrd/settings").json()
    assert g["email_subject"] == new_subj
    assert g["email_body"] == new_body


# ---------- Birth date on payslip: PUT ----------
def test_put_payslip_persists_tanggal_lahir(her_sess):
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    assert lst, "expected slips for 8/2026 already imported"
    slip = lst[0]
    sid = slip["id"]
    payload = {
        "period_month": slip["period_month"],
        "period_year": slip["period_year"],
        "employee_id": slip.get("employee_id"),
        "nik": slip.get("nik", ""),
        "nama": slip.get("nama", ""),
        "email": slip.get("email", ""),
        "jabatan": slip.get("jabatan", ""),
        "dept": slip.get("dept", "Production"),
        "no_rekening": slip.get("no_rekening", ""),
        "bank": slip.get("bank", ""),
        "earnings": [dict(e) for e in slip.get("earnings", [])],
        "deductions": [dict(e) for e in slip.get("deductions", [])],
        "take_home": slip.get("take_home"),
        "notes": slip.get("notes", ""),
        "tanggal_lahir": "1990-08-17",
    }
    r = her_sess.put(f"{API}/hrd/payslips/{sid}", json=payload)
    assert r.status_code == 200, r.text
    saved = r.json()
    assert saved.get("tanggal_lahir") == "1990-08-17"

    lst2 = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    s2 = next(p for p in lst2 if p["id"] == sid)
    assert s2.get("tanggal_lahir") == "1990-08-17"


# ---------- Importer auto-detects birth date from sheet ----------
def test_import_auto_detects_tanggal_lahir(her_sess):
    from openpyxl import load_workbook
    dst = "/tmp/gaji_lahir.xlsx"
    shutil.copyfile(XLSX_SRC, dst)
    wb = load_workbook(dst)
    # pick first slip sheet (A5 == 'SLIP GAJI')
    target = None
    for name in wb.sheetnames:
        ws = wb[name]
        if str(ws["A5"].value or "").strip().upper() == "SLIP GAJI":
            target = name
            break
    assert target, "no slip sheet found"
    ws = wb[target]
    # inject a 'Tgl Lahir : 1990-08-17' in an empty cell of row 11 col L (unused)
    ws["B11"] = "Tgl Lahir : 1990-08-17"
    wb.save(dst)

    with open(dst, "rb") as f:
        files = {"file": ("gaji.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = her_sess.post(f"{API}/hrd/payslips/import-excel",
                          data={"month": 10, "year": 2026},
                          files=files, timeout=60)
    assert r.status_code == 200, r.text
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 10, "year": 2026}).json()["items"]
    # find slip corresponding to the target sheet name (sheet name typically = nama)
    matched = [p for p in lst if target.lower() in (p.get("nama") or "").lower()]
    assert matched, f"no imported slip matches sheet {target}: {[p['nama'] for p in lst]}"
    assert matched[0].get("tanggal_lahir") == "1990-08-17", f"got {matched[0].get('tanggal_lahir')}"


def test_reimport_preserves_manual_tanggal_lahir(her_sess):
    # first set a manual tanggal_lahir on a slip in 10/2026
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 10, "year": 2026}).json()["items"]
    # pick one whose sheet DOES NOT have the lahir label; safest: pick one currently with empty tanggal_lahir
    victims = [p for p in lst if not p.get("tanggal_lahir")]
    assert victims, "expected at least one slip without tanggal_lahir"
    slip = victims[0]
    sid = slip["id"]
    payload = {
        "period_month": slip["period_month"], "period_year": slip["period_year"],
        "employee_id": slip.get("employee_id"),
        "nik": slip.get("nik", ""), "nama": slip.get("nama", ""),
        "email": slip.get("email", ""),
        "jabatan": slip.get("jabatan", ""), "dept": slip.get("dept", "Production"),
        "no_rekening": slip.get("no_rekening", ""), "bank": slip.get("bank", ""),
        "earnings": [dict(e) for e in slip.get("earnings", [])],
        "deductions": [dict(e) for e in slip.get("deductions", [])],
        "take_home": slip.get("take_home"), "notes": "",
        "tanggal_lahir": "1985-01-02",
    }
    r = her_sess.put(f"{API}/hrd/payslips/{sid}", json=payload)
    assert r.status_code == 200

    # re-import ORIGINAL xlsx (no lahir label) over month 10
    with open(XLSX_SRC, "rb") as f:
        files = {"file": ("gaji.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = her_sess.post(f"{API}/hrd/payslips/import-excel",
                          data={"month": 10, "year": 2026},
                          files=files, timeout=60)
    assert r.status_code == 200

    lst2 = her_sess.get(f"{API}/hrd/payslips", params={"month": 10, "year": 2026}).json()["items"]
    s2 = next(p for p in lst2 if p["id"] == sid)
    assert s2.get("tanggal_lahir") == "1985-01-02", f"manual tanggal_lahir wiped: {s2.get('tanggal_lahir')}"


# ---------- PDF encryption via internal render function ----------
def test_render_pdf_encrypted_with_birth_password():
    import sys
    sys.path.insert(0, "/app/backend")
    from routers import hrd as hrd_mod  # noqa
    slip = {
        "nama": "Test Encrypt", "nik": "999", "jabatan": "QA", "dept": "IT",
        "period_month": 8, "period_year": 2026,
        "earnings": [{"label": "Gaji Pokok", "qty": None, "amount": 1000000}],
        "deductions": [],
        "gross": 1000000, "total_deduction": 0, "net": 1000000, "take_home": 1000000,
        "terbilang": "satu juta", "tanggal_lahir": "1990-08-17",
    }
    pw = hrd_mod._birth_password(slip)
    assert pw == "17081990", f"password wrong: {pw}"
    buf = hrd_mod._render_slip_pdf(slip, password=pw)
    data = buf.read()
    assert data[:4] == b"%PDF"
    assert b"/Encrypt" in data, "PDF should be encrypted with password"

    buf2 = hrd_mod._render_slip_pdf(slip, password=None)
    data2 = buf2.read()
    assert data2[:4] == b"%PDF"
    assert b"/Encrypt" not in data2, "PDF preview should NOT be encrypted"


def test_pdf_preview_endpoint_returns_unencrypted(her_sess):
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    assert lst
    r = her_sess.get(f"{API}/hrd/payslips/{lst[0]['id']}/pdf")
    assert r.status_code == 200
    assert "application/pdf" in r.headers.get("content-type", "")
    assert r.content[:4] == b"%PDF"
    # Preview endpoint is unencrypted
    assert b"/Encrypt" not in r.content


# ---------- Blast without Gmail App Password ----------
def test_blast_400_when_no_app_password(her_sess):
    # ensure Gmail is not configured — clear via direct settings save is not possible (app_password only-set-when-provided)
    # so we assume the test env doesn't have it; if it does, we skip.
    s = her_sess.get(f"{API}/hrd/settings").json()
    if s.get("has_app_password"):
        pytest.skip("Gmail App Password is configured in this env; cannot verify 400 guard.")
    r = her_sess.post(f"{API}/hrd/blast", json={"month": 8, "year": 2026})
    assert r.status_code == 400, r.text
    body = r.json()
    detail = (body.get("detail") or "").lower()
    assert "gmail" in detail or "belum" in detail or "app password" in detail
