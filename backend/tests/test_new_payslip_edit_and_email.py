"""New feature tests (iteration 3):
- PUT /api/hrd/payslips/{id} updates identity, email, earnings, take_home
- Excel importer auto-detects email placed in any cell of the slip sheet
- Re-import preserves manually-set email when sheet has none
- Regression: month=8 year=2026 import creates/updates 5 slips, Wawan take_home=8,138,000
- 403s on GET /api/hrd/payslips for heri & susanto
- PDF returns application/pdf
"""
import io
import os
import shutil
import pytest
import requests

BASE = (os.environ.get("REACT_APP_BACKEND_URL") or "").rstrip("/")
if not BASE:
    with open("/app/frontend/.env") as f:
        for ln in f:
            if ln.startswith("REACT_APP_BACKEND_URL="):
                BASE = ln.split("=", 1)[1].strip().rstrip("/")
API = f"{BASE}/api"
XLSX_SRC = "/tmp/gaji.xlsx"

HER = ("herliana", "123456")
SU = ("susanto", "Subbalo1994")
HERI = ("heri", "123456")
GAJI_PIN = "5678"


def _login(u, p):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": u, "password": p}, timeout=15)
    assert r.status_code == 200, f"login {u} -> {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def her_sess():
    s = _login(*HER)
    v = s.post(f"{API}/hrd/verify-pin", json={"pin": GAJI_PIN})
    assert v.status_code == 200, f"gaji verify: {v.status_code} {v.text}"
    s.headers.update({"x-hrd-gaji": v.json()["gaji_token"]})
    return s


@pytest.fixture(scope="module")
def su_sess():
    return _login(*SU)


@pytest.fixture(scope="module")
def heri_sess():
    return _login(*HERI)


# ---------- Regression: month=8 year=2026 import & Wawan take_home ----------
def test_1_import_aug_2026(her_sess):
    with open(XLSX_SRC, "rb") as f:
        files = {"file": ("gaji.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = her_sess.post(f"{API}/hrd/payslips/import-excel",
                          data={"month": 8, "year": 2026},
                          files=files, timeout=60)
    assert r.status_code == 200, r.text
    d = r.json()
    assert len(d["sheets"]) == 5, f"sheets={d['sheets']}"
    assert (d["created"] + d["updated"]) == 5
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    wawan = [p for p in lst if "wawan" in (p.get("nama") or "").lower()]
    assert wawan, "wawan slip not found"
    assert int(wawan[0]["take_home"]) == 8_138_000, f"wawan take_home = {wawan[0]['take_home']}"


def test_2_pdf(her_sess):
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    assert lst
    r = her_sess.get(f"{API}/hrd/payslips/{lst[0]['id']}/pdf")
    assert r.status_code == 200
    assert "application/pdf" in r.headers.get("content-type", "")
    assert r.content[:4] == b"%PDF"


def test_3_forbidden_for_heri_and_susanto(heri_sess, su_sess):
    for s in (heri_sess, su_sess):
        r = s.get(f"{API}/hrd/payslips")
        assert r.status_code == 403


# ---------- PUT /payslips/{id}: identity + email + earning + take_home ----------
def test_4_update_payslip_full(her_sess):
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    wawan = next(p for p in lst if "wawan" in (p.get("nama") or "").lower())
    sid = wawan["id"]

    payload = {
        "period_month": wawan["period_month"],
        "period_year": wawan["period_year"],
        "employee_id": wawan.get("employee_id"),
        "nik": wawan.get("nik", ""),
        "nama": wawan.get("nama", ""),
        "email": "test1@mks.co.id",
        "jabatan": (wawan.get("jabatan") or "") + " (edit)",
        "dept": wawan.get("dept", "Production"),
        "no_rekening": wawan.get("no_rekening", ""),
        "bank": wawan.get("bank", ""),
        "earnings": [dict(e) for e in wawan.get("earnings", [])],
        "deductions": [dict(e) for e in wawan.get("deductions", [])],
        "take_home": wawan.get("take_home"),
        "notes": wawan.get("notes", ""),
    }
    # Change first earning by +1000
    if payload["earnings"]:
        payload["earnings"][0]["amount"] = float(payload["earnings"][0]["amount"]) + 1000

    r = her_sess.put(f"{API}/hrd/payslips/{sid}", json=payload)
    assert r.status_code == 200, r.text
    saved = r.json()
    assert saved["email"] == "test1@mks.co.id"
    assert saved["jabatan"].endswith("(edit)")
    assert "_id" not in saved

    # verify persisted via GET
    lst2 = her_sess.get(f"{API}/hrd/payslips", params={"month": 8, "year": 2026}).json()["items"]
    w2 = next(p for p in lst2 if p["id"] == sid)
    assert w2["email"] == "test1@mks.co.id"
    assert w2["jabatan"].endswith("(edit)")
    assert w2["gross"] == saved["gross"]


# ---------- Excel email auto-detection: new period 9/2026 with email injected ----------
def test_5_import_auto_detects_email_in_sheet(her_sess):
    from openpyxl import load_workbook
    dst = "/tmp/gaji_email.xlsx"
    shutil.copyfile(XLSX_SRC, dst)
    wb = load_workbook(dst)
    # Find wawan sheet (case-insensitive)
    wname = next(n for n in wb.sheetnames if n.lower() == "wawan")
    ws = wb[wname]
    ws["C11"] = "wawan@mks.co.id"
    wb.save(dst)

    with open(dst, "rb") as f:
        files = {"file": ("gaji.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = her_sess.post(f"{API}/hrd/payslips/import-excel",
                          data={"month": 9, "year": 2026},
                          files=files, timeout=60)
    assert r.status_code == 200, r.text
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 9, "year": 2026}).json()["items"]
    wawan = next(p for p in lst if "wawan" in (p.get("nama") or "").lower())
    assert wawan["email"] == "wawan@mks.co.id", f"got {wawan['email']}"


def test_6_reimport_preserves_manual_email(her_sess):
    """Re-import original xlsx (no sheet email) over slip that already has manual email
    must NOT wipe the email."""
    # First: for month=9/2026 slip, set an explicit email via PUT (simulate manual)
    lst = her_sess.get(f"{API}/hrd/payslips", params={"month": 9, "year": 2026}).json()["items"]
    # pick harjono to keep test independent from #5's wawan
    harj = next(p for p in lst if "harjono" in (p.get("nama") or "").lower())
    payload = {
        "period_month": harj["period_month"], "period_year": harj["period_year"],
        "employee_id": harj.get("employee_id"),
        "nik": harj.get("nik", ""), "nama": harj["nama"],
        "email": "manual.harjono@mks.co.id",
        "jabatan": harj.get("jabatan", ""), "dept": harj.get("dept", "Production"),
        "no_rekening": harj.get("no_rekening", ""), "bank": harj.get("bank", ""),
        "earnings": [dict(e) for e in harj.get("earnings", [])],
        "deductions": [dict(e) for e in harj.get("deductions", [])],
        "take_home": harj.get("take_home"), "notes": "",
    }
    r = her_sess.put(f"{API}/hrd/payslips/{harj['id']}", json=payload)
    assert r.status_code == 200

    # Now re-import the ORIGINAL xlsx (no email in sheets) over same period
    with open(XLSX_SRC, "rb") as f:
        files = {"file": ("gaji.xlsx", f,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = her_sess.post(f"{API}/hrd/payslips/import-excel",
                          data={"month": 9, "year": 2026},
                          files=files, timeout=60)
    assert r.status_code == 200, r.text

    lst2 = her_sess.get(f"{API}/hrd/payslips", params={"month": 9, "year": 2026}).json()["items"]
    h2 = next(p for p in lst2 if p["id"] == harj["id"])
    assert h2["email"] == "manual.harjono@mks.co.id", f"manual email was overwritten: got {h2['email']}"
