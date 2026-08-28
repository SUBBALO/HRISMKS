"""Payroll Excel import — EMAIL auto-detection robustness (bug fix verification).

Covers: plain email, uppercase+spaces, surrounding text ('Email: x@y'), trailing
non-breaking space, mailto hyperlink, name spacing mismatch (double space),
tolerant directory header ('E-Mail' / 'mail'), + regression of gaji_trial.xlsx
import, PDF 'T. Transport' and bulk-delete.
"""
import io
import os

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/")
API = f"{BASE_URL}/api"

MONTH, YEAR = 9, 2026
TRIAL_FILE = "/app/backend/tmp_uploads/gaji_trial.xlsx"
OUT_DIR = "/app/backend/tmp_uploads"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------- workbook builders ----------------
def _add_slip(wb, title, nama, nik, extra_cells=None, extra_links=None):
    ws = wb.create_sheet(title[:31])
    ws["A1"] = "PT MITRA KARYA SARANA"
    ws["A5"] = "SLIP GAJI"
    ws["B8"] = "Nama"; ws["C8"] = nama; ws["D8"] = "NIK"; ws["E8"] = nik
    ws["B9"] = "Dept"; ws["C9"] = "Production"
    ws["B10"] = "Jabatan"; ws["C10"] = "Operator"
    ws["G8"] = "Perhari"; ws["J8"] = 200000
    ws["G9"] = "Lembur/Jam"; ws["J9"] = 30000
    ws["G10"] = "T. Transport"; ws["J10"] = 50000
    for ref, val in (extra_cells or {}).items():
        ws[ref] = val
    if extra_links:
        from openpyxl.worksheet.hyperlink import Hyperlink
        for ref, target in extra_links.items():
            ws[ref] = ws[ref].value if ws[ref].value is not None else "klik"
            ws[ref].hyperlink = Hyperlink(ref=ref, target=target)
    r = 13
    for lbl, qty, amt in [("Gaji Pokok", None, 5000000), ("T. Tetap", None, 570000), ("Lembur (1.5)", 3, 90000)]:
        ws.cell(r, 1, lbl)
        if qty is not None:
            ws.cell(r, 3, qty)
        ws.cell(r, 5, amt)
        r += 1
    ws.cell(r, 1, "JUMLAH"); ws.cell(r, 5, 5660000)
    r = 13
    for lbl, amt in [("PPh 21", 91634), ("JHT+JP (2%+1%)", 169800)]:
        ws.cell(r, 7, lbl); ws.cell(r, 11, amt)
        r += 1
    ws.cell(r, 7, "JUMLAH"); ws.cell(r, 11, 261434)
    ws.cell(r + 1, 7, "PENGHASILAN BERSIH"); ws.cell(r + 1, 11, 5398566)
    ws.cell(r + 2, 7, "PEMBULATAN"); ws.cell(r + 2, 11, 5398600)
    return ws


# (dir_email raw value, expected email, slip nama override, extra slip cells/links)
CASES = [
    # 1) plain
    dict(nik="MKS 9101", dir_nama="TEST_Budi Santoso", slip_nama="TEST_Budi Santoso",
         dir_email="budi.qa@mks.co.id", expect="budi.qa@mks.co.id", label="plain"),
    # 2) uppercase + surrounding spaces
    dict(nik="MKS 9102", dir_nama="TEST_Ani Wijaya", slip_nama="TEST_Ani Wijaya",
         dir_email="  ANI.WIJAYA@MKS.CO.ID  ", expect="ani.wijaya@mks.co.id", label="uppercase+spaces"),
    # 3) surrounding text
    dict(nik="MKS 9103", dir_nama="TEST_Candra Kirana", slip_nama="TEST_Candra Kirana",
         dir_email="Email: candra@mks.co.id", expect="candra@mks.co.id", label="text-prefix"),
    # 4) trailing nbsp
    dict(nik="MKS 9104", dir_nama="TEST_Dewi Lestari", slip_nama="TEST_Dewi Lestari",
         dir_email="dewi@mks.co.id\u00a0", expect="dewi@mks.co.id", label="nbsp"),
    # 5) mailto hyperlink, cell text is not an email
    dict(nik="MKS 9105", dir_nama="TEST_Eko Prasetyo", slip_nama="TEST_Eko Prasetyo",
         dir_email="klik disini", dir_link="mailto:eko@mks.co.id",
         expect="eko@mks.co.id", label="dir-hyperlink"),
    # 6) name spacing mismatch (slip has double space)
    dict(nik="", dir_nama="TEST_Fajar Nugroho", slip_nama="TEST_Fajar  Nugroho",
         dir_email="fajar@mks.co.id", expect="fajar@mks.co.id", label="double-space-name"),
    # 7) email only inside slip sheet with surrounding text (no directory row)
    dict(nik="MKS 9107", dir_nama=None, slip_nama="TEST_Gita Sari",
         slip_cells={"B11": "Email: gita@mks.co.id"}, expect="gita@mks.co.id", label="slip-text-email"),
    # 8) email only as mailto hyperlink inside slip sheet
    dict(nik="MKS 9108", dir_nama=None, slip_nama="TEST_Hadi Wibowo",
         slip_cells={"B11": "kirim ke sini"}, slip_links={"B11": "mailto:hadi@mks.co.id"},
         expect="hadi@mks.co.id", label="slip-hyperlink"),
]


def _build_email_workbook(email_header="E-Mail"):
    from openpyxl import Workbook
    from openpyxl.worksheet.hyperlink import Hyperlink
    wb = Workbook()
    d = wb.active
    d.title = "Daftar Gaji"
    d["A1"] = "DAFTAR GAJI KARYAWAN"
    d["A3"] = "NIK"; d["B3"] = "Nama"; d["C3"] = email_header; d["D3"] = "Take Home Pay"
    r = 4
    for c in CASES:
        if not c.get("dir_nama"):
            continue
        d.cell(r, 1, c["nik"]); d.cell(r, 2, c["dir_nama"])
        cell = d.cell(r, 3, c["dir_email"])
        if c.get("dir_link"):
            cell.hyperlink = Hyperlink(ref=f"C{r}", target=c["dir_link"])
        d.cell(r, 4, 5398600)
        r += 1
    for i, c in enumerate(CASES):
        _add_slip(wb, f"Slip{i+1}", c["slip_nama"], c["nik"],
                  c.get("slip_cells"), c.get("slip_links"))
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    with open(os.path.join(OUT_DIR, f"TEST_email_variants_{email_header.replace('-','')}.xlsx"), "wb") as f:
        f.write(data)
    return data


# ---------------- fixtures ----------------
@pytest.fixture(scope="module")
def gaji_client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": "herliana", "password": "123456"}, timeout=30)
    if r.status_code != 200:
        pytest.fail(f"Login failed {r.status_code}: {r.text[:300]}")
    tok = r.json().get("access_token") or r.json().get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    p = s.post(f"{API}/hrd/verify-pin", json={"pin": "1234"}, timeout=30)
    if p.status_code != 200:
        pytest.fail(f"verify-pin failed {p.status_code}: {p.text[:300]}")
    token = p.json().get("gaji_token")
    assert token, f"no gaji_token in {p.text[:300]}"
    s.headers.update({"x-hrd-gaji": token})
    return s


def _list(client, month=MONTH, year=YEAR):
    r = client.get(f"{API}/hrd/payslips", params={"month": month, "year": year}, timeout=60)
    assert r.status_code == 200, f"{r.status_code}: {r.text[:300]}"
    payload = r.json()
    return payload if isinstance(payload, list) else payload.get("items", payload.get("data", []))


def _find(items, needle):
    return next((x for x in items if needle in str(x.get("nama", ""))), None)


# ---------------- tests ----------------
class TestEmailVariants:
    @pytest.fixture(scope="class", autouse=True)
    def imported(self, gaji_client):
        files = {"file": ("TEST_email_variants.xlsx", _build_email_workbook("E-Mail"), XLSX_MIME)}
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": MONTH, "year": YEAR}, files=files, timeout=180)
        assert r.status_code == 200, f"import failed {r.status_code}: {r.text[:600]}"
        d = r.json()
        assert d.get("success") is True, d
        assert (d.get("created", 0) + d.get("updated", 0)) == len(CASES), d
        return _list(gaji_client)

    @pytest.mark.parametrize("case", CASES, ids=[c["label"] for c in CASES])
    def test_email_detected(self, gaji_client, case):
        items = _list(gaji_client)
        key = case["slip_nama"].split()[0]  # e.g. TEST_Budi
        row = _find(items, key)
        assert row is not None, f"slip '{case['slip_nama']}' not imported"
        assert "_id" not in row, "MongoDB _id leaked"
        got = (row.get("email") or "").strip().lower()
        assert got, f"[{case['label']}] email EMPTY for {case['slip_nama']} (row={ {k: row.get(k) for k in ('nama','nik','email')} })"
        assert got == case["expect"], f"[{case['label']}] expected {case['expect']}, got {got!r}"

    def test_double_space_name_matched_directory(self, gaji_client):
        """Name in slip sheet has a double space vs directory — must still match."""
        row = _find(_list(gaji_client), "TEST_Fajar")
        assert row is not None
        assert (row.get("email") or "").lower() == "fajar@mks.co.id"
        # audit against 'Take Home Pay' column proves directory row matched
        assert row.get("dg_take_home") == 5398600, row.get("dg_take_home")


class TestDirectoryHeaderVariant:
    def test_header_mail_lowercase_variant(self, gaji_client):
        """Directory header spelled 'mail' must still be treated as the email column."""
        files = {"file": ("TEST_email_hdr.xlsx", _build_email_workbook("mail"), XLSX_MIME)}
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": MONTH, "year": YEAR}, files=files, timeout=180)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:600]}"
        items = _list(gaji_client)
        for case in CASES:
            row = _find(items, case["slip_nama"].split()[0])
            assert row is not None, case["slip_nama"]
            assert (row.get("email") or "").strip().lower() == case["expect"], \
                f"[header=mail][{case['label']}] got {row.get('email')!r}"


class TestRegressionTrialImport:
    def test_import_gaji_trial(self, gaji_client):
        if not os.path.exists(TRIAL_FILE):
            pytest.skip(f"missing {TRIAL_FILE}")
        with open(TRIAL_FILE, "rb") as f:
            content = f.read()
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": 8, "year": 2026},
                             files={"file": ("gaji_trial.xlsx", content, XLSX_MIME)}, timeout=180)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:600]}"
        d = r.json()
        assert d.get("success") is True, d
        assert (d.get("created", 0) + d.get("updated", 0)) == 3, d
        names = " | ".join(d.get("names", []))
        for n in ("Wawan", "Harjono", "Rahmat"):
            assert n in names, f"{n} missing in {names}"

    def test_pdf_rahmat_has_t_transport(self, gaji_client):
        items = _list(gaji_client, 8, 2026)
        row = _find(items, "Rahmat")
        assert row is not None, "Rahmat slip not found for 8/2026"
        r = gaji_client.get(f"{API}/hrd/payslips/{row['id']}/pdf", timeout=120)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:300]}"
        assert r.content[:4] == b"%PDF", "not a PDF"
        try:
            from pypdf import PdfReader
        except ImportError:
            try:
                from PyPDF2 import PdfReader
            except ImportError:
                pytest.skip("no pdf reader lib installed")
        text = "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(r.content)).pages)
        assert "Transport" in text, f"'T. Transport' missing in PDF text: {text[:800]}"


class TestBulkDeleteRegression:
    def test_bulk_delete_removes_slips(self, gaji_client):
        items = _list(gaji_client)
        ids = [x["id"] for x in items if str(x.get("nama", "")).startswith("TEST_")][:2]
        assert len(ids) >= 1, "no TEST_ slips available to bulk-delete"
        r = gaji_client.post(f"{API}/hrd/payslips/bulk-delete", json={"ids": ids}, timeout=60)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:300]}"
        assert r.json().get("deleted") == len(ids), r.json()
        remaining = {x["id"] for x in _list(gaji_client)}
        assert not (set(ids) & remaining), "bulk-deleted slips still listed"


@pytest.fixture(scope="module", autouse=True)
def cleanup(gaji_client):
    yield
    try:
        ids = [x["id"] for x in _list(gaji_client) if str(x.get("nama", "")).startswith("TEST_")]
        if ids:
            gaji_client.post(f"{API}/hrd/payslips/bulk-delete", json={"ids": ids}, timeout=60)
    except Exception as e:
        print(f"cleanup skipped: {e}")
