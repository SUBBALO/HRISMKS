"""Payroll Excel template download + Excel import tests (openpyxl fix verification)."""
import io
import os

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/")
API = f"{BASE_URL}/api"

MONTH, YEAR = 7, 2026


def _build_slip(nama="TEST_Budi Santoso", nik="MKS 9021"):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Slip TestBudi"
    ws["A1"] = "PT MITRA KARYA SARANA"
    ws["A5"] = "SLIP GAJI"
    ws["B8"] = "Nama"; ws["C8"] = nama; ws["D8"] = "NIK"; ws["E8"] = nik
    ws["B9"] = "Dept"; ws["C9"] = "Production"
    ws["B10"] = "Jabatan"; ws["C10"] = "Operator"
    ws["J8"] = 200000
    ws["J9"] = 3
    ws["J10"] = 50000
    r = 13
    for lbl, qty, amt in [("Gaji Pokok", None, 5000000), ("T. Tetap", None, 570000), ("Lembur (1.5)", 3, 90000)]:
        ws.cell(r, 1, lbl)
        if qty is not None:
            ws.cell(r, 3, qty)
        ws.cell(r, 5, amt)
        r += 1
    ws.cell(r, 1, "JUMLAH"); ws.cell(r, 5, 5660000)
    r = 13
    for lbl, amt in [("PPh 21", 91634), ("JHT+JP (2%+1%)", 169800), ("BPJS KESEHATAN 1%", 50000)]:
        ws.cell(r, 7, lbl); ws.cell(r, 11, amt)
        r += 1
    ws.cell(r, 7, "JUMLAH"); ws.cell(r, 11, 311434)
    ws.cell(r + 1, 7, "PENGHASILAN BERSIH"); ws.cell(r + 1, 11, 5348566)
    ws.cell(r + 2, 7, "PEMBULATAN"); ws.cell(r + 2, 11, 5348600)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_non_slip():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Random data"
    ws["B2"] = 123
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"username": "herliana", "password": "123456"}, timeout=30)
    if r.status_code != 200:
        pytest.fail(f"Login failed {r.status_code}: {r.text[:300]}")
    data = r.json()
    tok = data.get("access_token") or data.get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="module")
def gaji_client(client):
    r = client.post(f"{API}/hrd/verify-pin", json={"pin": "1234"}, timeout=30)
    if r.status_code != 200:
        pytest.fail(f"verify-pin failed {r.status_code}: {r.text[:300]}")
    token = r.json().get("gaji_token")
    assert token, f"no gaji_token in {r.text[:300]}"
    client.headers.update({"x-hrd-gaji": token})
    return client


# --- Template download ---
class TestImportTemplate:
    def test_template_download(self, gaji_client):
        r = gaji_client.get(f"{API}/hrd/import-template", timeout=60)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:400]}"
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or "excel" in ct, ct
        assert r.content[:2] == b"PK", "not a valid xlsx (zip) payload"
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), data_only=True)
        assert len(wb.sheetnames) >= 1

    def test_template_requires_pin(self, client):
        s = requests.Session()
        s.headers.update(dict(client.headers))
        s.headers.pop("x-hrd-gaji", None)
        s.cookies.update(client.cookies)
        r = s.get(f"{API}/hrd/import-template", timeout=30)
        assert r.status_code in (401, 403), f"expected pin gate, got {r.status_code}"


# --- Excel import ---
class TestImportExcel:
    def test_import_valid_slip_and_persist(self, gaji_client):
        files = {"file": ("TEST_slip.xlsx", _build_slip(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": MONTH, "year": YEAR}, files=files, timeout=120)
        assert r.status_code == 200, f"{r.status_code}: {r.text[:600]}"
        d = r.json()
        assert d.get("success") is True, d
        assert (d.get("created", 0) + d.get("updated", 0)) >= 1, d
        assert isinstance(d.get("names"), list) and any("TEST_Budi" in n for n in d["names"]), d

        lr = gaji_client.get(f"{API}/hrd/payslips", params={"month": MONTH, "year": YEAR}, timeout=60)
        assert lr.status_code == 200, lr.text[:300]
        payload = lr.json()
        items = payload if isinstance(payload, list) else payload.get("items", payload.get("data", []))
        row = next((x for x in items if "TEST_Budi" in str(x.get("nama", ""))), None)
        assert row is not None, f"imported slip not found in list ({len(items)} rows)"
        assert "_id" not in row, "MongoDB _id leaked in response"
        assert row.get("period_month") == MONTH and row.get("period_year") == YEAR
        assert float(row.get("gaji_pokok") or 0) == 5000000 or float(row.get("take_home") or 0) > 0, row

    def test_reimport_updates_not_duplicates(self, gaji_client):
        files = {"file": ("TEST_slip.xlsx", _build_slip(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": MONTH, "year": YEAR}, files=files, timeout=120)
        assert r.status_code == 200, r.text[:400]
        d = r.json()
        assert d.get("updated", 0) >= 1, f"expected upsert update, got {d}"

        lr = gaji_client.get(f"{API}/hrd/payslips", params={"month": MONTH, "year": YEAR}, timeout=60)
        payload = lr.json()
        items = payload if isinstance(payload, list) else payload.get("items", payload.get("data", []))
        matches = [x for x in items if "TEST_Budi" in str(x.get("nama", ""))]
        assert len(matches) == 1, f"duplicate rows created: {len(matches)}"

    def test_import_non_slip_xlsx_returns_400(self, gaji_client):
        files = {"file": ("TEST_bad.xlsx", _build_non_slip(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": MONTH, "year": YEAR}, files=files, timeout=60)
        assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text[:400]}"
        assert "SLIP GAJI" in r.text.upper() or "slip" in r.text.lower()

    def test_import_garbage_file_returns_400(self, gaji_client):
        files = {"file": ("TEST_bad.xlsx", b"not-an-excel-file", "application/octet-stream")}
        r = gaji_client.post(f"{API}/hrd/payslips/import-excel",
                             data={"month": MONTH, "year": YEAR}, files=files, timeout=60)
        assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text[:400]}"

    def test_import_without_pin_blocked(self, client):
        s = requests.Session()
        s.headers.update({k: v for k, v in client.headers.items() if k.lower() != "x-hrd-gaji"})
        s.headers.pop("x-hrd-gaji", None)
        s.cookies.update(client.cookies)
        files = {"file": ("TEST_slip.xlsx", _build_slip(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = s.post(f"{API}/hrd/payslips/import-excel",
                   data={"month": MONTH, "year": YEAR}, files=files, timeout=60)
        assert r.status_code in (401, 403), f"expected pin gate, got {r.status_code}"


@pytest.fixture(scope="module", autouse=True)
def cleanup(request):
    yield
    try:
        s = requests.Session()
        r = s.post(f"{API}/auth/login", json={"username": "herliana", "password": "123456"}, timeout=30)
        tok = r.json().get("access_token") or r.json().get("token")
        if tok:
            s.headers.update({"Authorization": f"Bearer {tok}"})
        pin = s.post(f"{API}/hrd/verify-pin", json={"pin": "1234"}, timeout=30).json().get("gaji_token")
        s.headers.update({"x-hrd-gaji": pin})
        lr = s.get(f"{API}/hrd/payslips", params={"month": MONTH, "year": YEAR}, timeout=60)
        payload = lr.json()
        items = payload if isinstance(payload, list) else payload.get("items", payload.get("data", []))
        for x in items:
            if "TEST_" in str(x.get("nama", "")):
                s.delete(f"{API}/hrd/payslips/{x.get('id')}", timeout=30)
    except Exception as e:
        print(f"cleanup skipped: {e}")
