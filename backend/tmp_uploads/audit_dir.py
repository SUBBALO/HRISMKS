import sys, os
sys.path.insert(0, "/app/backend")
os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "test_database")
os.environ.setdefault("JWT_SECRET", "x")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from routers import hrd

wb = load_workbook("/app/backend/tmp_uploads/gaji_trial.xlsx", data_only=True)
print("ALL SHEETS:", wb.sheetnames)
for sh in ["Daftar Gaji", "TER"]:
    if sh not in wb.sheetnames:
        continue
    ws = wb[sh]
    print("\n==== SHEET:", sh, "(A5=", repr(ws['A5'].value), ") ====")
    for r in range(1, 12):
        cells = []
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                cells.append(f"{get_column_letter(c)}{r}={v!r}")
        if cells:
            print("  " + " | ".join(cells))

dir_by_nama, dir_by_nik = hrd._parse_directory(wb)
print("\n==== DIRECTORY EMAIL per employee ====")
for nm in ["wawan munandar", "harjono", "rahmat ari sandi nst"]:
    rec = dir_by_nama.get(nm)
    print(f"  {nm:24s} -> email={rec.get('email')!r} nik={rec.get('nik')!r} th={rec.get('take_home')!r}" if rec else f"  {nm}: (none)")
print("\n==== JUNK / numeric-name keys in directory ====")
for k, v in dir_by_nama.items():
    if k not in ("wawan munandar", "harjono", "rahmat ari sandi nst"):
        print(f"  key={k!r} -> {v}")
