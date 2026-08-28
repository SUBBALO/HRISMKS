import sys, os
sys.path.insert(0, "/app/backend")
os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "test_database")
os.environ.setdefault("JWT_SECRET", "x")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from routers import hrd

wb = load_workbook("/app/backend/tmp_uploads/gaji_trial.xlsx", data_only=True)
slip_sheets = [n for n in wb.sheetnames if str(hrd._cell(wb[n], "A5") or "").strip().upper() == "SLIP GAJI"]
print("SLIP SHEETS:", slip_sheets)
dir_by_nama, dir_by_nik = hrd._parse_directory(wb)
print("DIR keys:", list(dir_by_nama.keys()))

for n in slip_sheets:
    ws = wb[n]
    print("\n" + "=" * 70)
    print("SHEET:", n)
    print("--- RAW GRID (rows 6-30, cols A-L) ---")
    for r in range(6, 31):
        cells = []
        for c in range(1, 13):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                cells.append(f"{get_column_letter(c)}{r}={v!r}")
        if cells:
            print("  " + " | ".join(cells))
    slip = hrd._parse_slip_sheet(ws, 7, 2026)
    print("--- PARSED ---")
    print("  nama    :", repr(slip.get("nama")))
    print("  nik     :", repr(slip.get("nik")))
    print("  dept    :", repr(slip.get("dept")), "| jabatan:", repr(slip.get("jabatan")))
    print("  rates   :", slip.get("rates"))
    print("  EARNINGS:")
    for e in slip["earnings"]:
        print("     ", e)
    print("  DEDUCTIONS:")
    for d in slip["deductions"]:
        print("     ", d)
    print("  gross   :", slip.get("gross"), "| total_ded:", slip.get("total_deduction"))
    print("  net     :", slip.get("net"), "| take_home:", slip.get("take_home"))
    print("  terbilang:", repr(slip.get("terbilang")))
    print("  email_sheet:", repr(slip.get("email_sheet")), "| tgl_lahir:", repr(slip.get("tanggal_lahir")))
