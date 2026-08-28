import sys, os
sys.path.insert(0, "/app/backend")
os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "test_database")
os.environ.setdefault("JWT_SECRET", "x")
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from routers import hrd

wb = Workbook()
# Directory sheet 'Daftar Gaji' with header variant 'E-Mail' and tricky values
d = wb.active
d.title = "Daftar Gaji"
d["A1"] = "DAFTAR GAJI KARYAWAN"
d["A3"] = "NIK"; d["B3"] = "Nama"; d["C3"] = "E-Mail"; d["D3"] = "Take Home Pay"
rows = [
    ("MKS001", "Budi Santoso", "budi@mks.co.id", 5000000),          # normal
    ("MKS002", "Ani Wijaya", "  ANI.WIJAYA@MKS.CO.ID  ", 4500000),  # uppercase + spaces
    ("MKS003", "Candra Kirana", "Email: candra@mks.co.id", 4200000),# surrounding text
    ("MKS004", "Dewi Lestari", "dewi@mks.co.id\u00a0", 4300000),     # trailing nbsp
    ("MKS005", "Eko Prasetyo", "", 4100000),                          # hyperlink only (set below)
    ("MKS006", "Budi  Santoso  Jr", "junior@mks.co.id", 3900000),    # double space in name
]
r = 4
for nik, nama, email, th in rows:
    d.cell(r, 1, nik); d.cell(r, 2, nama); d.cell(r, 3, email); d.cell(r, 4, th)
    r += 1
# Eko: email as a mailto hyperlink on an otherwise 'blank-looking' cell
cell = d.cell(8, 3, "click")
cell.hyperlink = Hyperlink(ref="C8", target="mailto:eko@mks.co.id")

dir_by_nama, dir_by_nik = hrd._parse_directory(wb)
print("=== directory parsed ===")
for nik, nama, *_ in rows:
    key = hrd._nk(nama)
    rec = dir_by_nama.get(key) or dir_by_nik.get(hrd._nk(nik))
    print(f"{nama:22s} -> {rec['email'] if rec else '(NO MATCH)'!r}")

# also test matching a slip whose name has different spacing than directory
print("\n=== slip-name matching ===")
for slip_name in ["budi santoso", "Budi  Santoso  Jr", "ANI WIJAYA"]:
    rec = dir_by_nama.get(hrd._nk(slip_name))
    print(f"slip '{slip_name}' -> {rec['email'] if rec else '(NO MATCH)'!r}")
