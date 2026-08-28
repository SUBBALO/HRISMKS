from openpyxl import Workbook

wb = Workbook(); ws = wb.active; ws.title = "Slip UITest"
ws["A1"] = "PT MITRA KARYA SARANA"; ws["A5"] = "SLIP GAJI"
ws["B8"] = "Nama"; ws["C8"] = "TEST_UI Rahmat"; ws["D8"] = "NIK"; ws["E8"] = "MKS 9099"
ws["B9"] = "Dept"; ws["C9"] = "Production"; ws["B10"] = "Jabatan"; ws["C10"] = "Operator"
ws["J8"] = 200000; ws["J9"] = 2; ws["J10"] = 50000
r = 13
for lbl, amt in [("Gaji Pokok", 4000000), ("T. Tetap", 500000)]:
    ws.cell(r, 1, lbl); ws.cell(r, 5, amt); r += 1
ws.cell(r, 1, "JUMLAH"); ws.cell(r, 5, 4500000)
r = 13
for lbl, amt in [("PPh 21", 50000), ("BPJS KESEHATAN 1%", 40000)]:
    ws.cell(r, 7, lbl); ws.cell(r, 11, amt); r += 1
ws.cell(r, 7, "JUMLAH"); ws.cell(r, 11, 90000)
ws.cell(r + 1, 7, "PENGHASILAN BERSIH"); ws.cell(r + 1, 11, 4410000)
ws.cell(r + 2, 7, "PEMBULATAN"); ws.cell(r + 2, 11, 4410000)
wb.save("/app/tests/fixtures/slip_ui.xlsx")

wb2 = Workbook(); wb2.active["A1"] = "junk data not a slip"; wb2.save("/app/tests/fixtures/bad_ui.xlsx")
print("ok")
